# -*- coding: utf-8 -*-
"""
校区学员情况管理 Skill - HTML报告生成脚本

功能：读取带标签汇总表.xlsx → 计算统计数据 → 序列化JSON → 注入HTML模板
     → 输出单文件校区分析报告.html（三页结构）

三页结构：
- 页1：全员表格+筛选（按校区/年级/支付力/风险筛选排序）+ 学情画像列
- 页2：校区分析报告（高净值占比/小区分布/学校层次分布/决策标签分布
       + 渠道来源分布/新签阶段漏斗/学情画像标签分布/流失风险预警/高净值待挖）
- 页3：产品推荐画像（推荐课程分布+百分比+画像筛选
       + 按学情画像分组推荐/按等级考路径推荐/按家长关注度推荐
       + 课程级精准推荐 v1.7.0新增）

v1.7.0：
- PROFILE_COURSE_MAP 从 utils.py 统一导入
- 学员列表新增"顾问侧续费历史"字段
- 课程产品库有数据时生成课程级精准推荐统计，无数据时降级为方向级推荐

用法：
    python scripts/generate_html_report.py --input <带标签汇总表.xlsx> --output <报告.html> --config <校区配置.json>

依赖：openpyxl>=3.1.2
"""

import argparse
import json
import os
import sys
from collections import Counter
from typing import Any, Dict, List

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from utils import (
    SUMMARY_COLUMNS, PROFILE_TAGS, PROFILE_TAG_PRIORITY,
    PROFILE_COURSE_MAP,
    SUBJECTIVE_FIELDS, EXCLUDED_FROM_UNCERTAINTY,
    match_course_products, generate_course_recommendation,
    get_timestamp, print_script_result,
)

# v1.7.0：PROFILE_COURSE_MAP 已移至 utils.py 作为唯一定义源，此处保留引用供本模块使用
# 学情画像标签→推荐课程方向映射（降级用，课程产品库为空或无匹配时使用）

# 等级考级别→推荐方向映射
EXAM_PATH_MAP = {
    "advanced": {"label": "中高级（已考3级以上）", "recommend": "竞赛冲刺/CSP/蓝桥杯"},
    "beginner": {"label": "初级（1-2级）", "recommend": "等级考进阶/Python进阶"},
    "none": {"label": "未考级", "recommend": "Scratch启蒙/等级考起步"},
}

# 家长关注度→推荐方向映射
ATTENTION_MAP = {
    "高": "竞赛路线/深度规划",
    "中": "进阶课程/考级规划",
    "低": "体验课激活/基础巩固",
}


def _classify_exam_level(exam_text: str) -> str:
    """将等级考文本分类为 advanced/beginner/none。

    Args:
        exam_text: 等级考字段文本，如"机器人3级"、"Python1级"、"未考"

    Returns:
        "advanced" / "beginner" / "none"
    """
    text = (exam_text or "").strip()
    if not text or text == "未考" or text == "无":
        return "none"
    # 提取数字
    import re
    nums = re.findall(r"\d+", text)
    if nums:
        level = int(nums[0])
        if level >= 3:
            return "advanced"
        else:
            return "beginner"
    return "none"


def _classify_attention(attention_text: str) -> str:
    """将家长关注度文本分类为 高/中/低。

    Args:
        attention_text: 家长关注度字段文本

    Returns:
        "高" / "中" / "低"
    """
    text = (attention_text or "").strip()
    if not text:
        return "中"
    # 群内回复快+经常进班=高
    if ("快" in text and "经常进" in text) or "高" in text:
        return "高"
    # 回复慢/不进班/疲态=低
    if ("慢" in text and "不进" in text) or "低" in text:
        return "低"
    if "疲态" in text and "疲态" not in text.replace("无疲态", ""):
        return "低"
    # 疲态:轻度/明显 → 低
    if ("轻度" in text or "明显" in text) and "疲态" in text:
        return "低"
    return "中"


def read_tagged_xlsx(file_path: str) -> List[Dict[str, Any]]:
    """读取带标签汇总表.xlsx，解析为学生记录列表。

    Args:
        file_path: 汇总表xlsx路径

    Returns:
        学生记录列表
    """
    from openpyxl import load_workbook

    wb = load_workbook(file_path, data_only=True)
    if "学员汇总" not in wb.sheetnames:
        print(f"[错误] 汇总表中未找到'学员汇总'sheet")
        return []

    ws = wb["学员汇总"]
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value) if cell.value else "")

    students: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue

        student: Dict[str, Any] = {}
        for idx, header in enumerate(headers):
            val = str(row[idx]) if idx < len(row) and row[idx] is not None else ""
            student[header] = val
        students.append(student)

    return students


def parse_priority_to_int(priority_str: str) -> int:
    """将跟进优先级字符串解析为整数（⭐符号计数）。

    Args:
        priority_str: 优先级字符串，如"⭐⭐⭐⭐"或"4"

    Returns:
        星级数（1-5）
    """
    if not priority_str:
        return 0
    # 数字符串中的⭐数量
    star_count = priority_str.count("⭐")
    if star_count > 0:
        return star_count
    # 尝试解析数字
    try:
        return int(priority_str)
    except ValueError:
        return 0


def build_report_data(students: List[Dict[str, Any]], campus: str,
                      course_library: List[Dict[str, Any]] = None) -> Dict[str, Any]:
    """构建HTML报告数据结构。

    Args:
        students: 学生记录列表
        campus: 校区名
        course_library: 校区课程产品库（v1.7.0），为空时降级为方向级推荐

    Returns:
        报告数据字典（含学员列表和统计信息）
    """
    if course_library is None:
        course_library = []
    total = len(students)

    # 构建学员列表（精简字段用于前端展示）+ 新增B2/D2/B_cross字段
    student_list: List[Dict[str, Any]] = []
    for s in students:
        student_list.append({
            "姓名": s.get("姓名", ""),
            "年级": s.get("年级", ""),
            "校区": s.get("校区", campus),
            "居住小区": s.get("居住小区", ""),
            "学校层次": s.get("学校层次(科技特色)", ""),
            "支付力": s.get("支付力", ""),
            "续费风险": s.get("续费风险", ""),
            "转介绍潜力": s.get("转介绍潜力", ""),
            "跟进优先级": parse_priority_to_int(s.get("跟进优先级", "")),
            "推荐方向": s.get("推荐产品方向", ""),
            "推荐理由": s.get("推荐理由", "") if "推荐理由" in s else "",
            # v1.8 责任关系与筛选标签字段
            "当前顾问": s.get("当前顾问", ""),
            "历史顾问": s.get("历史顾问", ""),
            "交接状态": s.get("交接状态", ""),
            "归属老师标签": s.get("归属老师标签", ""),
            "课程段标签": s.get("课程段标签", ""),
            "年龄段标签": s.get("年龄段标签", ""),
            # B2销售漏斗字段
            "客户来源": s.get("客户来源", ""),
            "当前阶段": s.get("当前阶段", ""),
            "顾问侧续费历史": s.get("顾问侧续费历史", ""),
            # D2学情履历字段
            "学情画像": s.get("学情画像", ""),
            "入学时间": s.get("入学时间", ""),
            "在读时长": s.get("在读时长", ""),
            "等级考": s.get("等级考", ""),
            "白名单比赛": s.get("白名单比赛", ""),
            "老师侧支付力": s.get("老师侧支付力", ""),
            "家长关注度": s.get("家长关注度", ""),
            # v1.8.1 新增字段
            "可信度标记": s.get("可信度标记", ""),
            "年龄": s.get("年龄", ""),
        })

    # 统计计算
    payment_counter = Counter(s.get("支付力", "") for s in students)
    risk_counter = Counter(s.get("续费风险", "") for s in students)
    referral_counter = Counter(s.get("转介绍潜力", "") for s in students)
    recommend_counter = Counter(s.get("推荐产品方向", "") for s in students)
    course_segment_counter = Counter((s.get("课程段标签", "") or "未填写").strip() or "未填写" for s in students)
    teacher_owner_counter = Counter((s.get("归属老师标签", "") or "未填写").strip() or "未填写" for s in students)
    advisor_counter = Counter((s.get("当前顾问", "") or "未填写").strip() or "未填写" for s in students)
    handoff_counter = Counter((s.get("交接状态", "") or "未填写").strip() or "未填写" for s in students)
    age_segment_counter = Counter((s.get("年龄段标签", "") or "未填写").strip() or "未填写" for s in students)

    # 高净值占比
    high_payment_count = payment_counter.get("高", 0)
    high_payment_pct = f"{(high_payment_count / total * 100):.0f}%" if total > 0 else "0%"

    # 续费高风险占比
    high_risk_count = risk_counter.get("高", 0)
    high_risk_pct = f"{(high_risk_count / total * 100):.0f}%" if total > 0 else "0%"

    # 高转介绍潜力占比
    high_referral_count = referral_counter.get("高", 0)
    high_referral_pct = f"{(high_referral_count / total * 100):.0f}%" if total > 0 else "0%"

    # 家长来源小区分布
    community_counter = Counter()
    for s in students:
        community = s.get("居住小区", "").strip()
        if community and community != "不知道" and community != "不清楚":
            community_counter[community] += 1
    community_dist = [{"小区": k, "人数": v} for k, v in community_counter.most_common(10)]

    # 学校层次分布
    school_counter = Counter()
    for s in students:
        school_level = s.get("学校层次(科技特色)", "").strip()
        if not school_level:
            school_level = "未补齐"
        if "科技特色" in school_level:
            school_level = "科技特色校"
        elif "重点" in school_level:
            school_level = "重点校"
        elif "普通" in school_level:
            school_level = "普通校"
        school_counter[school_level] += 1
    school_dist = [{"层次": k, "人数": v} for k, v in school_counter.most_common()]

    # 推荐分布
    recommend_dist = []
    for course, count in recommend_counter.most_common():
        pct = f"{(count / total * 100):.0f}%" if total > 0 else "0%"
        recommend_dist.append({"课程": course if course else "待配置", "人数": count, "占比": pct})

    # 决策标签分布
    tag_dist = []
    for level in ["高", "中", "低"]:
        tag_dist.append({"标签": f"支付力-{level}", "人数": payment_counter.get(level, 0)})
        tag_dist.append({"标签": f"续费风险-{level}", "人数": risk_counter.get(level, 0)})
        tag_dist.append({"标签": f"转介绍潜力-{level}", "人数": referral_counter.get(level, 0)})

    # ===== 新增统计（页2）=====

    # 0. v1.8.0 责任关系与筛选标签摘要
    course_segment_dist = [{"课程段": k, "人数": v} for k, v in course_segment_counter.most_common()]
    teacher_owner_dist = [{"老师": k, "人数": v} for k, v in teacher_owner_counter.most_common()]
    advisor_dist = [{"顾问": k, "人数": v} for k, v in advisor_counter.most_common()]
    handoff_dist = [{"交接状态": k, "人数": v} for k, v in handoff_counter.most_common()]
    age_segment_dist = [{"年龄段": k, "人数": v} for k, v in age_segment_counter.most_common()]

    # 1. 渠道来源分布（客户来源 Counter）
    channel_counter = Counter()
    for s in students:
        channel = (s.get("客户来源", "") or "").strip()
        if channel:
            channel_counter[channel] += 1
    channel_dist = [{"渠道": k, "人数": v} for k, v in channel_counter.most_common()]

    # 2. 新签阶段漏斗（当前阶段 Counter，按固定顺序排序）
    stage_order = ["新签挖需", "诺访", "在读更新"]
    stage_counter = Counter()
    for s in students:
        stage = (s.get("当前阶段", "") or "").strip()
        if stage:
            stage_counter[stage] += 1
    funnel_dist = []
    for stage in stage_order:
        if stage_counter.get(stage, 0) > 0:
            funnel_dist.append({"阶段": stage, "人数": stage_counter[stage]})
    # 追加不在stage_order中的其他阶段
    for stage, count in stage_counter.most_common():
        if stage not in stage_order:
            funnel_dist.append({"阶段": stage, "人数": count})

    # 3. 学情画像标签分布（遍历学员学情画像列，多标签用"、"分隔，按PROFILE_TAGS顺序输出）
    profile_tag_counter = Counter()
    for s in students:
        profile_text = (s.get("学情画像", "") or "").strip()
        if profile_text:
            tags = [t.strip() for t in profile_text.split("、") if t.strip()]
            for tag in tags:
                profile_tag_counter[tag] += 1
    profile_dist = []
    for tag in PROFILE_TAGS:
        count = profile_tag_counter.get(tag, 0)
        if count > 0:
            profile_dist.append({"标签": tag, "人数": count})

    # 4. 流失风险预警列表（学情画像含"流失风险型"）
    risk_warning_list = []
    for s in students:
        profile_text = (s.get("学情画像", "") or "").strip()
        if "流失风险型" in profile_text:
            risk_warning_list.append({
                "姓名": s.get("姓名", ""),
                "年级": s.get("年级", ""),
                "校区": s.get("校区", campus),
                "学情画像": profile_text,
            })

    # 5. 高净值待挖列表（学情画像含"高净值待挖型"）
    potential_list = []
    for s in students:
        profile_text = (s.get("学情画像", "") or "").strip()
        if "高净值待挖型" in profile_text:
            potential_list.append({
                "姓名": s.get("姓名", ""),
                "年级": s.get("年级", ""),
                "校区": s.get("校区", campus),
                "学情画像": profile_text,
                "老师侧支付力": s.get("老师侧支付力", ""),
            })

    # ===== 新增推荐分组（页3）=====

    # 1. 按学情画像分组推荐
    profile_recommend = []
    profile_groups: Dict[str, List[str]] = {}
    for s in students:
        profile_text = (s.get("学情画像", "") or "").strip()
        if profile_text:
            tags = [t.strip() for t in profile_text.split("、") if t.strip()]
            for tag in tags:
                if tag not in profile_groups:
                    profile_groups[tag] = []
                profile_groups[tag].append(s.get("姓名", ""))
    # 按PROFILE_TAG_PRIORITY顺序输出
    for tag in PROFILE_TAGS:
        if tag in profile_groups and profile_groups[tag]:
            recommend_course = PROFILE_COURSE_MAP.get(tag, "待配置")
            profile_recommend.append({
                "学情画像": tag,
                "推荐课程方向": recommend_course,
                "学员数": len(profile_groups[tag]),
                "学员名单": "、".join(profile_groups[tag]),
            })

    # 2. 按等级考路径推荐
    exam_path_groups: Dict[str, List[str]] = {"advanced": [], "beginner": [], "none": []}
    for s in students:
        exam_text = s.get("等级考", "")
        level = _classify_exam_level(exam_text)
        exam_path_groups[level].append(s.get("姓名", ""))
    exam_path_recommend = []
    for level in ["advanced", "beginner", "none"]:
        info = EXAM_PATH_MAP[level]
        exam_path_recommend.append({
            "等级考级别": info["label"],
            "推荐方向": info["recommend"],
            "学员数": len(exam_path_groups[level]),
        })

    # 3. 按家长关注度推荐
    attention_groups: Dict[str, List[str]] = {"高": [], "中": [], "低": []}
    for s in students:
        attention_text = s.get("家长关注度", "")
        level = _classify_attention(attention_text)
        attention_groups[level].append(s.get("姓名", ""))
    attention_recommend = []
    for level in ["高", "中", "低"]:
        attention_recommend.append({
            "家长关注度": level,
            "推荐方向": ATTENTION_MAP[level],
            "学员数": len(attention_groups[level]),
        })

    # 4. 课程级精准推荐统计（v1.7.0新增）
    #    课程产品库有数据时，对每个学员调用 generate_course_recommendation，
    #    统计每门课程被推荐了几次、推荐了哪些学员
    course_level_recommend: List[Dict[str, Any]] = []
    if course_library:
        # 课程名→推荐学员名单 的映射
        course_student_map: Dict[str, List[str]] = {}
        # 课程名→课程信息 的映射（从课程库取核心卖点、开班数）
        course_info_map: Dict[str, Dict[str, Any]] = {}
        for course in course_library:
            cname = course.get("课程名", "")
            if cname:
                course_info_map[cname] = course
                course_student_map.setdefault(cname, [])

        # 遍历学员，调用 generate_course_recommendation
        for s in students:
            rec_result = generate_course_recommendation(s, course_library, PROFILE_COURSE_MAP)
            if rec_result.get("推荐类型") == "course_level":
                matched_courses = rec_result.get("推荐课程", [])
                for mc in matched_courses:
                    cname = mc.get("课程名", "")
                    if cname and cname in course_student_map:
                        student_name = s.get("姓名", "")
                        if student_name and student_name not in course_student_map[cname]:
                            course_student_map[cname].append(student_name)

        # 构建统计列表（仅包含有推荐学员的课程）
        for course in course_library:
            cname = course.get("课程名", "")
            if not cname:
                continue
            recommend_students = course_student_map.get(cname, [])
            if not recommend_students:
                continue
            selling_points = course.get("核心卖点", [])
            if not isinstance(selling_points, list):
                selling_points = [selling_points] if selling_points else []
            course_level_recommend.append({
                "课程名": cname,
                "核心卖点": selling_points,
                "当前开班数": course.get("当前开班数", 0),
                "推荐学员数": len(recommend_students),
                "推荐学员名单": recommend_students,
            })

        # 按推荐学员数降序排序
        course_level_recommend.sort(key=lambda x: -x["推荐学员数"])

    # ===== v1.8.1 新增统计：数据质量与责任全景 =====

    # 主观字段完成率（遍历SUBJECTIVE_FIELDS对应的主表列，统计非空比）
    subjective_total = len(SUBJECTIVE_FIELDS) * total if total > 0 else 1
    subjective_non_empty = 0
    for s in students:
        for field in SUBJECTIVE_FIELDS:
            val = (s.get(field, "") or "").strip()
            if val:
                subjective_non_empty += 1
    subjective_completion_rate = f"{(subjective_non_empty / subjective_total * 100):.0f}%" if subjective_total > 0 else "0%"
    subjective_completion_float = round(subjective_non_empty / subjective_total * 100, 1) if subjective_total > 0 else 0.0

    # 客观字段完成率（A基础标识+D1课程成果+D2客观字段）
    objective_fields = ["校区", "姓名", "年级"]  # A基础标识（去掉年龄因为已移除）
    objective_fields.extend(["已报名课程", "在读课程", "学习时长", "作品成果", "续费历史"])  # D1
    objective_fields.extend(["入学时间", "在读时长", "等级考", "白名单比赛", "老师侧支付力",
                              "家长关注度", "家长新期待", "老师复盘"])  # D2
    obj_total = len(objective_fields) * total if total > 0 else 1
    obj_non_empty = 0
    for s in students:
        for field in objective_fields:
            val = (s.get(field, "") or "").strip()
            if val:
                obj_non_empty += 1
    objective_completion_rate = f"{(obj_non_empty / obj_total * 100):.0f}%" if obj_total > 0 else "0%"
    objective_completion_float = round(obj_non_empty / obj_total * 100, 1) if obj_total > 0 else 0.0

    # 可信度分布 Counter
    credibility_counter = Counter((s.get("可信度标记", "") or "未标记").strip() or "未标记" for s in students)
    credibility_dist = [
        {"可信度": level, "人数": credibility_counter.get(level, 0)}
        for level in ["高可信度", "中可信度", "低可信度", "待验证", "未标记"]
    ]

    # 冲突统计
    conflict_students = [s for s in students if (s.get("冲突标注", "") or "").strip()]
    conflict_field_count = 0
    for s in conflict_students:
        conflict_text = (s.get("冲突标注", "") or "").strip()
        # 格式如 "冲突字段: 成绩水平; 课堂表现"
        if "冲突字段:" in conflict_text:
            fields_part = conflict_text.replace("冲突字段:", "").strip()
            conflict_field_count += len([f for f in fields_part.split(";") if f.strip()])
        else:
            conflict_field_count += 1
    conflict_stats = {
        "冲突字段总数": conflict_field_count,
        "涉及学生数": len(conflict_students),
        "冲突学生列表": [
            {"姓名": s.get("姓名", ""), "年级": s.get("年级", ""),
             "冲突标注": s.get("冲突标注", "")}
            for s in conflict_students[:20]  # 最多20条
        ],
    }

    # 待确认汇总列表（可信度=待验证 或 低可信度 的学生）
    pending_list = []
    for s in students:
        cred = (s.get("可信度标记", "") or "").strip()
        if cred in ("待验证", "低可信度"):
            pending_list.append({
                "姓名": s.get("姓名", ""),
                "年级": s.get("年级", ""),
                "校区": s.get("校区", campus),
                "可信度标记": cred,
                "冲突标注": s.get("冲突标注", ""),
            })

    # 字段完成率明细（逐字段非空占比，含所有SUMMARY_COLUMNS中的信息字段）
    field_completion_detail = []
    info_columns = [c for c in SUMMARY_COLUMNS if c not in (
        "校区", "姓名", "年级", "冲突标注", "可信度标记",
        "学校层次(科技特色)", "科技特色详情", "小区房价段", "住户画像",
        "周边竞品", "家庭消费力", "推荐话术素材",
        "支付力", "续费风险", "转介绍潜力", "跟进优先级", "推荐产品方向", "学情画像",
        "年龄段标签", "交接状态",
    )]
    for field in info_columns:
        non_empty = sum(1 for s in students if (s.get(field, "") or "").strip())
        rate = f"{(non_empty / total * 100):.0f}%" if total > 0 else "0%"
        field_completion_detail.append({"字段": field, "非空人数": non_empty, "完成率": rate})

    # 主观字段完成率明细（用于柱状图）
    subjective_field_detail = []
    for field in SUBJECTIVE_FIELDS:
        non_empty = sum(1 for s in students if (s.get(field, "") or "").strip())
        rate = round(non_empty / total * 100, 1) if total > 0 else 0.0
        subjective_field_detail.append({"字段": field, "非空人数": non_empty, "完成率": rate})

    # 组装报告数据
    report_data = {
        "生成时间": get_timestamp(),
        "校区": campus,
        "学员列表": student_list,
        "统计": {
            "总人数": total,
            "高净值占比": high_payment_pct,
            "续费高风险占比": high_risk_pct,
            "高转介绍潜力占比": high_referral_pct,
            "家长来源小区分布": community_dist,
            "学校层次分布": school_dist,
            "推荐分布": recommend_dist,
            "决策标签分布": tag_dist,
            # 页2新增统计
            "课程段分布": course_segment_dist,
            "老师归属分布": teacher_owner_dist,
            "顾问分布": advisor_dist,
            "交接状态分布": handoff_dist,
            "年龄段分布": age_segment_dist,
            "渠道来源分布": channel_dist,
            "新签阶段漏斗": funnel_dist,
            "学情画像标签分布": profile_dist,
            "流失风险预警": risk_warning_list,
            "高净值待挖列表": potential_list,
            # 页3新增推荐分组
            "按学情画像分组推荐": profile_recommend,
            "按等级考路径推荐": exam_path_recommend,
            "按家长关注度推荐": attention_recommend,
            # v1.7.0新增：课程级精准推荐（课程产品库有数据时，否则为空列表→降级方向级）
            "课程级推荐": course_level_recommend,
            # v1.8.1 新增：数据质量与责任全景（页4/页5）
            "主观字段完成率": subjective_completion_rate,
            "主观字段完成率数值": subjective_completion_float,
            "客观字段完成率": objective_completion_rate,
            "客观字段完成率数值": objective_completion_float,
            "可信度分布": credibility_dist,
            "冲突统计": conflict_stats,
            "待确认汇总": pending_list,
            "字段完成率明细": field_completion_detail,
            "主观字段完成率明细": subjective_field_detail,
        }
    }

    return report_data


def inject_data_to_html(template_path: str, report_data: Dict[str, Any],
                        output_path: str) -> bool:
    """将报告数据注入HTML模板，生成单文件报告。

    Args:
        template_path: HTML模板文件路径
        report_data: 报告数据字典
        output_path: 输出HTML文件路径

    Returns:
        生成成功返回True
    """
    # 读取模板
    with open(template_path, "r", encoding="utf-8") as f:
        html_content = f.read()

    # 将数据序列化为JSON
    data_json = json.dumps(report_data, ensure_ascii=False, indent=2)

    # 替换模板中的REPORT_DATA占位
    # 模板中有: const REPORT_DATA = { ... };
    # 我们用实际数据替换，使用贪婪匹配处理嵌套花括号
    import re
    pattern = r'const REPORT_DATA\s*=\s*\{.*?\};'
    replacement = f'const REPORT_DATA = {data_json};'
    html_content = re.sub(pattern, replacement, html_content, count=1, flags=re.DOTALL)

    # 写入输出文件
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True) if os.path.dirname(output_path) else None
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html_content)

    return True


def main():
    """主函数：解析参数并生成HTML报告。"""
    parser = argparse.ArgumentParser(description="生成HTML单文件报告（三页结构）")
    parser.add_argument("--input", required=True, help="带标签汇总表xlsx路径")
    parser.add_argument("--output", required=True, help="输出HTML报告路径")
    parser.add_argument("--config", help="校区配置.json路径（读取报告偏好，可选）")
    args = parser.parse_args()

    # 检查文件
    if not os.path.exists(args.input):
        print_script_result(False, f"汇总表不存在: {args.input}")
        sys.exit(1)

    # 读取配置（可选）
    campus = ""
    course_library: List[Dict[str, Any]] = []
    if args.config and os.path.exists(args.config):
        with open(args.config, "r", encoding="utf-8") as f:
            config = json.load(f)
        campus = config.get("校区", "")
        # v1.7.0：读取课程产品库（可选，为空时降级为方向级推荐）
        course_library = config.get("课程产品库", [])
        if not isinstance(course_library, list):
            course_library = []

    # 读取汇总表
    students = read_tagged_xlsx(args.input)
    if not students:
        print_script_result(False, "汇总表为空或读取失败")
        sys.exit(1)

    # 如果未从配置获取校区，从数据中取
    if not campus and students:
        campus = students[0].get("校区", "未知校区")

    # 构建报告数据
    report_data = build_report_data(students, campus, course_library)

    # 查找HTML模板
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(script_dir)
    template_path = os.path.join(project_root, "templates", "html_report_template.html")

    if not os.path.exists(template_path):
        print_script_result(False, f"HTML模板不存在: {template_path}")
        sys.exit(1)

    # 注入数据生成报告
    try:
        inject_data_to_html(template_path, report_data, args.output)
        print_script_result(
            True,
            f"HTML报告生成成功：{args.output}",
            校区=campus,
            总人数=report_data["统计"]["总人数"],
            高净值占比=report_data["统计"]["高净值占比"],
            续费高风险占比=report_data["统计"]["续费高风险占比"],
        )
    except Exception as e:
        print_script_result(False, f"HTML报告生成异常: {str(e)}")
        sys.exit(1)


if __name__ == "__main__":
    main()
