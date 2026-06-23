# -*- coding: utf-8 -*-
"""
校区学员情况管理 Skill - 公共工具函数模块

提供字段映射、冲突格式化、日期处理、Excel样式等公共功能。
被 export_student_xlsx.py / merge_xlsx.py / write_tags.py / generate_html_report.py 共用。

依赖：openpyxl>=3.1.2
"""

import json
import os
import re
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

# ============================================================
# 字段定义：三层架构字段清单
# ============================================================

# A. 基础标识字段（顾问+老师共录）
A_FIELDS: List[str] = ["姓名", "昵称", "年级", "年龄", "所在校区"]

# B. 家庭背景与规划字段（顾问主录）
B_FIELDS: List[str] = ["家长职业", "单位性质", "家庭结构", "教育氛围",
                       "居住小区", "家长规划目标", "家长教育取向", "家长竞赛认知", "对AI认知度"]

# C. 学生在校情况字段（老师主录）
C_FIELDS: List[str] = ["学校名称", "成绩水平", "性格特点",
                       "兴趣偏好", "课堂表现", "同伴关系"]

# D. 在读课程与成果字段（老师主录）
D_FIELDS: List[str] = ["已报名课程", "在读课程", "学习时长",
                       "作品成果", "续费历史"]

# E. 学员细节备注字段（顾问+老师均可补充）
E_FIELDS: List[str] = ["学员细节备注"]

# B2. 销售漏斗字段（顾问主录，v1.7.0新增B2.09，共9字段）
B2_FIELDS: List[str] = [
    "客户来源",         # B2.01 选择题：大众点评/小红书/转介绍/公众号/地推/其他 ✅必填
    "对接次数",         # B2.02 数字 ✅必填
    "累计跟进时长",     # B2.03 文本"3周""2个月" ✅必填
    "当前阶段",         # B2.04 选择题：新签挖需/诺访/在读更新 ✅必填
    "最初兴趣点",       # B2.05 文本
    "介绍过的产品",     # B2.06 文本
    "堵点",             # B2.07 主观题
    "顾问复盘",         # B2.08 AI生成+顾问确认，200字内
    "顾问侧续费历史",   # B2.09 v1.7.0新增，文本
]

# D2. 学情履历字段（老师主录，新增13字段）
D2_FIELDS: List[str] = [
    "入学时间",       # D2.01 日期YYYY-MM ✅必填
    "入学时年级",     # D2.02 String ✅必填
    "当前年级",       # D2.03 自动计算
    "在读时长",       # D2.04 自动计算
    "过往奖项履历",   # D2.05 文本
    "特长兴趣",       # D2.06 文本
    "等级考",         # D2.07 多选，配置化
    "白名单比赛",     # D2.08 文本：比赛名+时间
    "老师侧支付力",   # D2.09 选择题：研学/lab/VEX等
    "学生状态观察",   # D2.10 自由文本（心情/性格/课堂表现合并）
    "家长关注度",     # D2.11 多维：群内回复+进班观察+疲态
    "家长新期待",     # D2.12 文本
    "老师复盘",       # D2.13 AI生成+老师确认，200字内
]

# 三路信源字段（家庭背景交叉补充）
CROSS_FIELDS: List[str] = ["家庭背景(老师补充)"]  # B_cross_teacher

# G. 责任关系与筛选标签字段（v1.8.0新增，主表当前快照）
RELATION_MAIN_FIELDS: List[str] = [
    "当前顾问", "历史顾问", "顾问关系备注", "交接状态",
    "归属老师标签", "课程段标签", "年龄段标签",
]

# 责任关系明细sheet列（v1.8.0新增）
RESPONSIBILITY_DETAIL_COLUMNS: List[str] = [
    "姓名", "年级", "关系类型", "关系状态", "负责人姓名",
    "课程段", "关系备注", "来源角色", "来源文件", "更新时间",
]

# 标准课程段值，前5项为课程段标签展示优先级；Python段用于无具体级别时保留原始关系语义。
COURSE_SEGMENT_VALUES: List[str] = ["Code2", "Code3", "PYTHON1", "PYTHON2", "PYTHON3", "Python段"]

# 课程段别名归一化配置。键为标准值，值为常见口语/大小写/中英文混合表达。
COURSE_SEGMENT_ALIASES: Dict[str, List[str]] = {
    "Code2": ["code2", "code 2", "code二", "code 二", "c2", "c 2", "代码二", "编程二"],
    "Code3": ["code3", "code 3", "code三", "code 三", "c3", "c 3", "代码三", "编程三"],
    "PYTHON1": ["python1", "python 1", "python一", "python 一", "py1", "py 1", "py一", "py 一", "p1", "p 1"],
    "PYTHON2": ["python2", "python 2", "python二", "python 二", "py2", "py 2", "py二", "py 二", "p2", "p 2"],
    "PYTHON3": ["python3", "python 3", "python三", "python 三", "py3", "py 3", "py三", "py 三", "p3", "p 3"],
    "Python段": ["python段", "python", "py段", "py", "python课程", "python老师"],
}

# 年龄段规则说明（infer_age_segment为唯一执行函数）
AGE_SEGMENT_RULES: List[Dict[str, str]] = [
    {"条件": "年龄<=8", "年龄段": "低龄段"},
    {"条件": "9<=年龄<=10", "年龄段": "小学中段"},
    {"条件": "11<=年龄<=12", "年龄段": "小学高段"},
    {"条件": "年龄>=13", "年龄段": "初中段及以上"},
    {"条件": "年龄缺失时按年级估算", "年龄段": "追加(按年级估算,待确认)"},
]

# 学情画像标签（8类，v1文字标签）
PROFILE_TAGS: List[str] = [
    "竞赛冲刺型", "科创潜力型", "兴趣探索型", "续费稳定型",
    "流失风险型", "高净值待挖型", "谨慎观望型", "基础夯实型",
]

# 学情画像优先级（命中≥3个时按此顺序取前2个为主标签）
PROFILE_TAG_PRIORITY: List[str] = [
    "流失风险型", "高净值待挖型", "竞赛冲刺型",
    "谨慎观望型", "续费稳定型", "科创潜力型",
    "兴趣探索型", "基础夯实型",
]

# 学情画像标签→推荐课程方向映射（降级用，课程产品库为空或无匹配时使用）
# v1.7.0：从 generate_html_report.py 移至 utils.py 作为唯一定义源，避免重复维护
PROFILE_COURSE_MAP: Dict[str, str] = {
    "竞赛冲刺型": "C++/信奥课程",
    "高净值待挖型": "研学/VEX",
    "兴趣探索型": "Scratch/机器人入门",
    "科创潜力型": "机器人进阶/科创项目",
    "续费稳定型": "进阶课程/续费",
    "流失风险型": "体验课重新激活",
    "谨慎观望型": "深度解答/试听",
    "基础夯实型": "等级考培训",
}

# AI补齐层字段
AI_FIELDS: List[str] = ["小区房价段", "住户画像", "周边学校",
                        "学校层次", "科技特色详情", "周边竞品",
                        "家庭消费力评估", "推荐话术素材"]

# 决策标签字段
TAG_FIELDS: List[str] = ["支付力", "续费风险", "转介绍潜力",
                         "跟进优先级", "推荐产品方向"]

# 顾问版Excel列顺序（A + B + B2 + 顾问关系快照 + E）= 5+9+9+4+1 = 28列
CONSULTANT_COLUMNS: List[str] = A_FIELDS + B_FIELDS + B2_FIELDS + [
    "当前顾问", "历史顾问", "顾问关系备注", "交接状态",
] + E_FIELDS

# 老师版Excel列顺序（A + C + D + D2 + E + B_cross_teacher + 老师标签）= 5+6+5+13+1+1+3 = 34列
TEACHER_COLUMNS: List[str] = A_FIELDS + C_FIELDS + D_FIELDS + D2_FIELDS + E_FIELDS + CROSS_FIELDS + [
    "归属老师标签", "课程段标签", "年龄段标签",
]

# 汇总表主表列顺序（全部字段，v1.8.1共59列）
# v1.8.1变更：删除"年龄"（已有年龄段标签替代）、合并"家长职业"+"单位性质"→"家长职业与单位"、
#           删除"对AI认知度"（保留在个人B表）、删除"同伴关系"（保留在个人C表）、
#           新增"可信度标记"列
SUMMARY_COLUMNS: List[str] = [
    # A基础标识（3列，v1.8.1移除"年龄"）
    "校区", "姓名", "年级",
    # B1家庭背景（7列；v1.8.1合并家长职业+单位性质→家长职业与单位，移除对AI认知度）
    "家长职业与单位", "家庭结构", "教育氛围", "居住小区",
    "家长规划目标", "家长教育取向", "家长竞赛认知",
    # B2销售漏斗（7列进汇总；最初兴趣点/介绍过的产品不进汇总；v1.7.0新增顾问侧续费历史）
    "客户来源", "对接次数", "累计跟进时长", "当前阶段", "堵点", "顾问复盘",
    "顾问侧续费历史",  # B2.09 v1.7.0新增
    # C在校情况（5列；v1.8.1移除"同伴关系"）
    "学校名称", "成绩水平", "性格特点", "兴趣偏好", "课堂表现",
    # D1课程成果（5列）
    "已报名课程", "在读课程", "学习时长", "作品成果", "续费历史",
    # D2学情履历（8列进汇总；入学时年级/当前年级/过往奖项/特长兴趣/学生状态观察不进汇总）
    "入学时间", "在读时长", "等级考", "白名单比赛", "老师侧支付力",
    "家长关注度", "家长新期待", "老师复盘",
    # E学员细节备注（1列）
    "学员细节备注",
    # 三路信源（1列）
    "家庭背景(老师补充)",
    # G责任关系与筛选标签（7列，主表当前快照）
    "当前顾问", "历史顾问", "顾问关系备注", "交接状态",
    "归属老师标签", "课程段标签", "年龄段标签",
    # 合并生成（1列）
    "冲突标注",
    # 可信度标记（1列，v1.8.1新增）
    "可信度标记",
    # AI补齐（7列）
    "学校层次(科技特色)", "科技特色详情", "小区房价段", "住户画像", "周边竞品", "家庭消费力",
    "推荐话术素材",
    # 决策标签（5列）
    "支付力", "续费风险", "转介绍潜力", "跟进优先级", "推荐产品方向",
    # 学情画像（1列）
    "学情画像",
]

# v1.8.1 可信度标记相关常量
# 主观字段列表（用于判断信息完整性）
SUBJECTIVE_FIELDS: List[str] = [
    "家长职业与单位", "家庭结构", "教育氛围", "居住小区",
    "家长规划目标", "家长教育取向", "家长竞赛认知",
    "学校名称", "成绩水平", "性格特点", "兴趣偏好", "课堂表现",
]
# 不确定表述关键词（匹配任一即判定为"低可信度"）
UNCERTAINTY_MARKERS: List[str] = [
    "大概", "可能", "不清楚", "不确定", "也许", "好像", "估计", "似乎",
]
# 排除字段（不参与不确定关键词扫描）
EXCLUDED_FROM_UNCERTAINTY: List[str] = [
    "年龄段标签", "交接状态", "冲突标注",
]


def calculate_credibility(student: Dict[str, Any], flat: Dict[str, str]) -> str:
    """计算学生记录的可信度标记。

    优先级由高到低：
    1. 扫描 flat 所有字段值（排除 EXCLUDED_FROM_UNCERTAINTY 字段），
       匹配 UNCERTAINTY_MARKERS 中任一词 → 返回"低可信度"
    2. flat["冲突标注"] 非空 → 返回"待验证"
    3. 所有 SUBJECTIVE_FIELDS 在 flat 中均为空 → 返回"待验证"
    4. student["来源角色"] 同时含"顾问"和"老师"且无冲突 → 返回"高可信度"
    5. 默认 → 返回"中可信度"

    Args:
        student: 合并后的学生记录（含"来源角色"、"冲突标注"等字段）
        flat: 扁平化后的字段字典

    Returns:
        "高可信度" / "中可信度" / "低可信度" / "待验证"
    """
    # 优先级1：扫描不确定关键词
    for key, value in flat.items():
        if key in EXCLUDED_FROM_UNCERTAINTY:
            continue
        val_str = str(value) if value else ""
        for marker in UNCERTAINTY_MARKERS:
            if marker in val_str:
                return "低可信度"

    # 优先级2：有冲突标注 → 待验证
    conflict = (flat.get("冲突标注", "") or "").strip()
    if conflict:
        return "待验证"

    # 优先级3：所有主观字段均为空 → 待验证
    all_empty = True
    for field in SUBJECTIVE_FIELDS:
        val = (flat.get(field, "") or "").strip()
        if val:
            all_empty = False
            break
    if all_empty:
        return "待验证"

    # 优先级4：来源角色同时含"顾问"和"老师"且无冲突 → 高可信度
    source_roles = student.get("来源角色", [])
    if isinstance(source_roles, list) and "顾问" in source_roles and "老师" in source_roles:
        return "高可信度"

    # 优先级5：默认
    return "中可信度"


# 冲突清单sheet列
CONFLICT_COLUMNS: List[str] = ["姓名", "年级", "字段名", "顾问值", "老师值", "状态", "备注"]

# 变更记录sheet列
CHANGELOG_COLUMNS: List[str] = ["姓名", "变更时间", "变更字段", "旧值", "新值", "变更来源"]

# 采集完成率sheet列
COMPLETION_COLUMNS: List[str] = ["校区", "采集人", "角色", "名单总数", "已采数", "完成率"]


# ============================================================
# 字段分组映射：字段名 → 所属分组
# ============================================================

FIELD_GROUP_MAP: Dict[str, str] = {}
for f in A_FIELDS:
    FIELD_GROUP_MAP[f] = "A基础标识"
for f in B_FIELDS:
    FIELD_GROUP_MAP[f] = "B1家庭背景"
for f in B2_FIELDS:
    FIELD_GROUP_MAP[f] = "B2销售漏斗"
for f in C_FIELDS:
    FIELD_GROUP_MAP[f] = "C在校情况"
for f in D_FIELDS:
    FIELD_GROUP_MAP[f] = "D1课程成果"
for f in D2_FIELDS:
    FIELD_GROUP_MAP[f] = "D2学情履历"
for f in CROSS_FIELDS:
    FIELD_GROUP_MAP[f] = "三路信源"
for f in RELATION_MAIN_FIELDS:
    FIELD_GROUP_MAP[f] = "G责任关系与筛选标签"


# ============================================================
# v1.8.0 责任关系与筛选标签函数
# ============================================================

def _clean_text(value: Any) -> str:
    """将任意值安全转成去首尾空白的字符串。

    Args:
        value: 任意输入值。

    Returns:
        规整后的字符串；None返回空字符串。
    """
    if value is None:
        return ""
    return str(value).strip()


def _dedupe_keep_order(values: List[str]) -> List[str]:
    """按出现顺序去重并丢弃空值。

    Args:
        values: 字符串列表。

    Returns:
        去重后的非空字符串列表。
    """
    seen = set()
    result: List[str] = []
    for value in values:
        clean_value = _clean_text(value)
        if not clean_value or clean_value in seen:
            continue
        seen.add(clean_value)
        result.append(clean_value)
    return result


def _chinese_digit_to_int(text: str) -> str:
    """将常见中文数字替换为阿拉伯数字字符。

    Args:
        text: 原始文本。

    Returns:
        替换后的文本。
    """
    compound_mapping = {
        "十二": "12", "十一": "11", "十": "10",
    }
    digit_mapping = {
        "零": "0", "一": "1", "二": "2", "两": "2", "三": "3",
        "四": "4", "五": "5", "六": "6", "七": "7", "八": "8", "九": "9",
    }
    result = text
    for cn_num, digit in compound_mapping.items():
        result = result.replace(cn_num, digit)
    for cn_num, digit in digit_mapping.items():
        result = result.replace(cn_num, digit)
    return result


def normalize_course_segment(raw: Any) -> str:
    """归一化课程段表达。

    支持 code2/Code 2/code二/C2、py1/Python一 等常见表达。
    无法识别时返回去空白后的原始字符串，避免丢失人工录入信息。

    Args:
        raw: 原始课程段文本。

    Returns:
        标准课程段值：Code2/Code3/PYTHON1/PYTHON2/PYTHON3/Python段；无法识别时返回原文。
    """
    text = _clean_text(raw)
    if not text:
        return ""

    compact = re.sub(r"[\s_\-　]+", "", text.lower())
    compact = _chinese_digit_to_int(compact)

    normalized_aliases: Dict[str, List[str]] = {}
    for standard, aliases in COURSE_SEGMENT_ALIASES.items():
        normalized_aliases[standard] = [
            _chinese_digit_to_int(re.sub(r"[\s_\-　]+", "", alias.lower()))
            for alias in aliases
        ]

    # 优先匹配具体课程段，避免"python"先命中Python段。
    for standard in COURSE_SEGMENT_VALUES:
        if standard == "Python段":
            continue
        standard_key = _chinese_digit_to_int(re.sub(r"[\s_\-　]+", "", standard.lower()))
        candidates = [standard_key] + normalized_aliases.get(standard, [])
        if compact in candidates or any(candidate and candidate in compact for candidate in candidates):
            return standard

    python_segment_candidates = normalized_aliases.get("Python段", [])
    if compact in python_segment_candidates:
        return "Python段"

    return text


def infer_age_segment(age: Any, grade: Any) -> str:
    """推断年龄段标签。

    年龄优先：≤8低龄段；9-10小学中段；11-12小学高段；≥13初中段及以上。
    年龄缺失时按年级兜底并追加"(按年级估算,待确认)"。

    Args:
        age: 年龄文本或数字。
        grade: 年级文本。

    Returns:
        年龄段标签；信息不足返回空字符串。
    """
    age_text = _clean_text(age)
    grade_text = _clean_text(grade)

    age_match = re.search(r"\d+", age_text)
    if age_match:
        age_value = int(age_match.group(0))
        if age_value <= 8:
            return "低龄段"
        if 9 <= age_value <= 10:
            return "小学中段"
        if 11 <= age_value <= 12:
            return "小学高段"
        return "初中段及以上"

    if not grade_text:
        return ""

    grade_normalized = _chinese_digit_to_int(grade_text)
    grade_match = re.search(r"\d+", grade_normalized)
    if not grade_match:
        return ""

    grade_value = int(grade_match.group(0))
    is_middle_school = any(keyword in grade_text for keyword in ["初", "七", "八", "九"]) or grade_value >= 7
    if is_middle_school:
        segment = "初中段及以上"
    elif grade_value <= 2:
        segment = "低龄段"
    elif grade_value <= 4:
        segment = "小学中段"
    elif grade_value <= 6:
        segment = "小学高段"
    else:
        segment = "初中段及以上"
    return f"{segment}(按年级估算,待确认)"


def normalize_relation_status(raw: Any) -> str:
    """归一化责任关系状态。

    Args:
        raw: 原始关系状态文本。

    Returns:
        当前/历史/待确认/已交接/顾问冲突待确认/空字符串/原文。
    """
    text = _clean_text(raw)
    if not text:
        return ""
    lowered = text.lower()
    if any(keyword in text for keyword in ["当前", "现在", "现任", "负责中", "正在负责"]):
        return "当前"
    if any(keyword in text for keyword in ["历史", "以前", "之前", "原来", "曾经", "过往"]):
        return "历史"
    if any(keyword in text for keyword in ["待确认", "确认中", "不确定", "未知"]):
        return "待确认"
    if "交接" in text and not any(keyword in text for keyword in ["未", "无", "没有"]):
        return "已交接"
    if "冲突" in text:
        return "顾问冲突待确认"
    if lowered in {"current", "active"}:
        return "当前"
    if lowered in {"history", "historical", "past", "previous"}:
        return "历史"
    return text


def _is_consultant_relation(relation: Dict[str, Any]) -> bool:
    """判断明细是否为顾问关系。"""
    relation_type = _clean_text(relation.get("关系类型"))
    if relation_type:
        return "顾问" in relation_type
    return any(_clean_text(relation.get(field)) for field in ["当前顾问", "历史顾问", "顾问关系备注"])


def _is_teacher_relation(relation: Dict[str, Any]) -> bool:
    """判断明细是否为老师关系。"""
    relation_type = _clean_text(relation.get("关系类型"))
    if relation_type:
        return "老师" in relation_type
    return bool(_clean_text(relation.get("归属老师标签")) or _clean_text(relation.get("课程段")))


def _get_relation_owner(relation: Dict[str, Any]) -> str:
    """读取责任关系负责人姓名，兼容多种缓存键名。"""
    for key in ["负责人姓名", "负责人", "老师", "顾问", "当前顾问", "历史顾问"]:
        value = _clean_text(relation.get(key))
        if value:
            return value
    return ""


def _split_multi_value(text: Any) -> List[str]:
    """拆分顿号/逗号/分号/竖线连接的多值文本。"""
    value = _clean_text(text)
    if not value:
        return []
    return [part.strip() for part in re.split(r"[、,，;；|｜/]+", value) if part.strip()]


def determine_handoff_status(relations: List[Dict[str, Any]]) -> str:
    """按顾问关系明细判断交接状态。

    规则：只有当前顾问=无交接；只有历史顾问=待确认；历史+当前不同=已交接；
    多个当前顾问冲突=顾问冲突待确认；全缺失为空。

    Args:
        relations: 责任关系明细列表。

    Returns:
        交接状态字符串。
    """
    current_advisors: List[str] = []
    historical_advisors: List[str] = []

    for relation in relations or []:
        if not isinstance(relation, dict) or not _is_consultant_relation(relation):
            continue
        status = normalize_relation_status(relation.get("关系状态"))
        owner = _get_relation_owner(relation)
        current_value = _clean_text(relation.get("当前顾问"))
        history_value = _clean_text(relation.get("历史顾问"))

        if current_value:
            current_advisors.extend(_split_multi_value(current_value))
        if history_value:
            historical_advisors.extend(_split_multi_value(history_value))
        if owner:
            if status == "当前":
                current_advisors.append(owner)
            elif status in {"历史", "已交接"}:
                historical_advisors.append(owner)
            elif status == "待确认" and not current_value:
                historical_advisors.append(owner)

    current_unique = _dedupe_keep_order(current_advisors)
    history_unique = _dedupe_keep_order(historical_advisors)

    if len(current_unique) > 1:
        return "顾问冲突待确认"
    if current_unique and not history_unique:
        return "无交接"
    if history_unique and not current_unique:
        return "待确认"
    if current_unique and history_unique:
        if any(history != current_unique[0] for history in history_unique):
            return "已交接"
        return "无交接"
    return ""


def build_teacher_tags(relations: List[Dict[str, Any]]) -> str:
    """从老师责任关系明细构建归属老师标签。

    Args:
        relations: 责任关系明细列表。

    Returns:
        形如"Code2-王老师、PYTHON1-李老师"的标签字符串。
    """
    tags: List[str] = []
    for relation in relations or []:
        if not isinstance(relation, dict) or not _is_teacher_relation(relation):
            continue
        explicit_tag = _clean_text(relation.get("归属老师标签"))
        if explicit_tag:
            tags.extend(_split_multi_value(explicit_tag))
            continue
        owner = _get_relation_owner(relation)
        segment = normalize_course_segment(relation.get("课程段"))
        if owner and segment:
            tags.append(f"{segment}-{owner}")
        elif owner:
            tags.append(owner)
    return "、".join(_dedupe_keep_order(tags))


def build_course_segment_tags(relations: List[Dict[str, Any]], course_text: Any = "") -> str:
    """从责任关系与课程文本构建课程段标签。

    Args:
        relations: 责任关系明细列表。
        course_text: 已报名/在读课程等文本，用于无明细时轻量补算。

    Returns:
        按 Code2、Code3、PYTHON1、PYTHON2、PYTHON3 优先级排序的课程段标签文本。
    """
    segments: List[str] = []
    for relation in relations or []:
        if not isinstance(relation, dict):
            continue
        segment = normalize_course_segment(relation.get("课程段"))
        if segment:
            segments.append(segment)

    raw_course = _clean_text(course_text)
    if raw_course:
        for part in re.split(r"[、,，;；|｜/\s]+", raw_course):
            normalized_part = normalize_course_segment(part)
            if normalized_part and normalized_part != part:
                segments.append(normalized_part)

        compact_course = _chinese_digit_to_int(re.sub(r"[\s_\-　]+", "", raw_course.lower()))
        for standard, aliases in COURSE_SEGMENT_ALIASES.items():
            if standard == "Python段":
                continue
            for alias in aliases + [standard]:
                compact_alias = _chinese_digit_to_int(re.sub(r"[\s_\-　]+", "", alias.lower()))
                if compact_alias and compact_alias in compact_course:
                    segments.append(standard)
                    break

    unique_segments = _dedupe_keep_order(segments)
    priority = {segment: index for index, segment in enumerate(COURSE_SEGMENT_VALUES)}
    unique_segments.sort(key=lambda item: priority.get(item, len(priority)))
    return "、".join(unique_segments)


def build_relation_snapshot(relations: List[Dict[str, Any]],
                            student: Optional[Dict[str, Any]] = None) -> Dict[str, str]:
    """构建主表责任关系与筛选标签快照。

    Args:
        relations: 责任关系明细列表。
        student: 学生扁平或嵌套记录，用于课程段/年龄段兜底推算。

    Returns:
        含RELATION_MAIN_FIELDS所有字段的字典。
    """
    current_advisors: List[str] = []
    historical_advisors: List[str] = []
    advisor_notes: List[str] = []
    safe_student = student or {}

    for relation in relations or []:
        if not isinstance(relation, dict) or not _is_consultant_relation(relation):
            continue
        status = normalize_relation_status(relation.get("关系状态"))
        owner = _get_relation_owner(relation)
        current_value = _clean_text(relation.get("当前顾问"))
        history_value = _clean_text(relation.get("历史顾问"))
        note = _clean_text(relation.get("关系备注", relation.get("顾问关系备注", "")))

        if current_value:
            current_advisors.extend(_split_multi_value(current_value))
        if history_value:
            historical_advisors.extend(_split_multi_value(history_value))
        if owner:
            if status == "当前":
                current_advisors.append(owner)
            elif status in {"历史", "已交接"}:
                historical_advisors.append(owner)
            elif status == "待确认" and not current_value:
                historical_advisors.append(owner)
        if note:
            advisor_notes.append(note)

    course_text_parts = [
        safe_student.get("在读课程", ""),
        safe_student.get("已报名课程", ""),
        safe_student.get("推荐产品方向", ""),
    ]
    course_group = safe_student.get("课程成果", {})
    if isinstance(course_group, dict):
        course_text_parts.extend([course_group.get("在读课程", ""), course_group.get("已报名课程", "")])
    course_text = "、".join(_clean_text(part) for part in course_text_parts if _clean_text(part))

    snapshot = {field: "" for field in RELATION_MAIN_FIELDS}
    snapshot["当前顾问"] = "、".join(_dedupe_keep_order(current_advisors))
    snapshot["历史顾问"] = "、".join(_dedupe_keep_order(historical_advisors))
    snapshot["顾问关系备注"] = " | ".join(_dedupe_keep_order(advisor_notes))
    snapshot["交接状态"] = determine_handoff_status(relations)
    snapshot["归属老师标签"] = build_teacher_tags(relations)
    snapshot["课程段标签"] = build_course_segment_tags(relations, course_text)
    snapshot["年龄段标签"] = infer_age_segment(safe_student.get("年龄", ""), safe_student.get("年级", ""))
    return snapshot


# ============================================================
# JSON缓存相关函数
# ============================================================

def get_timestamp() -> str:
    """获取当前时间的ISO 8601格式字符串。

    Returns:
        ISO 8601格式时间戳，如 "2026-06-22T10:30:00"
    """
    return datetime.now().strftime("%Y-%m-%dT%H:%M:%S")


def load_json(file_path: str) -> Dict[str, Any]:
    """加载JSON文件。

    Args:
        file_path: JSON文件路径

    Returns:
        解析后的字典；文件不存在或解析失败时返回空字典
    """
    if not os.path.exists(file_path):
        return {}
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError) as e:
        print(f"[警告] 加载JSON失败 {file_path}: {e}")
        return {}


def save_json(file_path: str, data: Dict[str, Any]) -> bool:
    """保存数据为JSON文件（UTF-8编码，缩进2空格）。

    Args:
        file_path: 输出文件路径
        data: 要保存的字典数据

    Returns:
        保存成功返回True，失败返回False
    """
    try:
        os.makedirs(os.path.dirname(file_path), exist_ok=True) if os.path.dirname(file_path) else None
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return True
    except IOError as e:
        print(f"[错误] 保存JSON失败 {file_path}: {e}")
        return False


def create_empty_student_record(name: str, campus: str, role: str) -> Dict[str, Any]:
    """创建空的学员档案记录。

    Args:
        name: 学生姓名
        campus: 所在校区
        role: 采集人角色（顾问/老师）

    Returns:
        包含所有字段（空值）的学员档案字典
    """
    record: Dict[str, Any] = {
        "姓名": name,
        "昵称": "",
        "年级": "",
        "年龄": "",
        "所在校区": campus,
        "家庭背景": {f: "" for f in B_FIELDS},          # B1（原有）
        "销售漏斗": {f: "" for f in B2_FIELDS},          # B2（新增）
        "在校情况": {f: "" for f in C_FIELDS},
        "课程成果": {f: "" for f in D_FIELDS},            # D1（原有）
        "学情履历": {f: "" for f in D2_FIELDS},           # D2（新增）
        "家庭背景_老师补充": "",                           # B_cross_teacher（新增）
        "责任关系": [],                                      # G责任关系明细（v1.8.0新增）
        "学员细节备注": "",
        "AI补齐": None,
        "决策标签": None,
        "采集人角色": role,
        "采集时间": "",
        "采集状态": "未采",
    }
    return record


def create_collection_cache(campus: str, role: str, collector_name: str) -> Dict[str, Any]:
    """创建空的采集端缓存结构。

    Args:
        campus: 校区名
        role: 采集人角色
        collector_name: 采集人姓名

    Returns:
        采集端缓存字典
    """
    return {
        "$schema": "collection_cache_v1",
        "校区": campus,
        "采集人角色": role,
        "采集人姓名": collector_name,
        "名单": [],
        "已采集记录": [],
        "最后更新时间": get_timestamp(),
    }


# ============================================================
# 冲突处理函数
# ============================================================

def format_conflict_cell(consultant_value: str, teacher_value: str) -> str:
    """格式化冲突字段为Excel单元格内双值格式。

    A+C双保留方案：同字段不同来源都保留，标注来源。

    Args:
        consultant_value: 顾问录入的值
        teacher_value: 老师录入的值

    Returns:
        格式化后的字符串，如 "顾问认为:xxx / 老师反馈:yyy [待核实]"
    """
    c_val = consultant_value.strip() if consultant_value else ""
    t_val = teacher_value.strip() if teacher_value else ""
    return f"顾问认为:{c_val} / 老师反馈:{t_val} [待核实]"


def create_conflict_record(name: str, grade: str, field: str,
                           consultant_value: str, teacher_value: str) -> Dict[str, Any]:
    """创建冲突记录。

    Args:
        name: 学生姓名
        grade: 年级
        field: 冲突字段名
        consultant_value: 顾问值
        teacher_value: 老师值

    Returns:
        冲突记录字典
    """
    return {
        "姓名": name,
        "年级": grade,
        "字段名": field,
        "顾问值": consultant_value,
        "老师值": teacher_value,
        "状态": "待核实",
        "备注": "",
    }


def detect_conflict(val1: str, val2: str) -> bool:
    """检测两个字段值是否冲突。

    两个值都非空且不相等时判定为冲突。

    Args:
        val1: 第一个值
        val2: 第二个值

    Returns:
        冲突返回True，否则False
    """
    v1 = (val1 or "").strip()
    v2 = (val2 or "").strip()
    if not v1 or not v2:
        return False
    return v1 != v2


# ============================================================
# 完成率计算函数
# ============================================================

def calculate_completion(cache: Dict[str, Any]) -> Dict[str, Any]:
    """计算采集完成率。

    Args:
        cache: 采集端缓存字典

    Returns:
        包含名单总数、已采数、完成率的字典
    """
    total = len(cache.get("名单", []))
    collected = len(cache.get("已采集记录", []))
    rate = f"{collected}/{total} ({(collected / total * 100):.0f}%)" if total > 0 else "0/0 (0%)"
    return {
        "名单总数": total,
        "已采数": collected,
        "完成率": rate,
    }


# ============================================================
# 姓名匹配函数
# ============================================================

def get_match_key(name: str, grade: str) -> Tuple[str, str]:
    """生成学生匹配键（姓名+年级）。

    唯一标识=中文正式名（完整含姓），重名用年级对齐。

    Args:
        name: 学生姓名
        grade: 年级

    Returns:
        (姓名, 年级) 元组作为匹配键
    """
    return (name.strip() if name else "", grade.strip() if grade else "")


def find_student_in_list(records: List[Dict[str, Any]],
                         name: str, grade: str) -> Optional[Dict[str, Any]]:
    """在记录列表中按姓名+年级查找学生。

    Args:
        records: 学生记录列表
        name: 要查找的姓名
        grade: 要查找的年级

    Returns:
        找到的记录字典，未找到返回None
    """
    target_key = get_match_key(name, grade)
    for record in records:
        rec_key = get_match_key(
            record.get("姓名", ""),
            record.get("年级", ""),
        )
        if rec_key == target_key:
            return record
    return None


# ============================================================
# Excel样式函数（依赖openpyxl）
# ============================================================

def get_excel_styles():
    """获取Excel样式对象。

    延迟导入openpyxl，避免未安装时模块加载失败。

    Returns:
        包含header_fill, header_font, border等样式对象的字典
    """
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("[错误] 未安装openpyxl，请执行: pip install openpyxl>=3.1.2")
        raise

    # 表头样式：加粗 + 浅蓝背景 + 居中
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 数据样式
    data_font = Font(name="微软雅黑", size=10)
    data_align = Alignment(vertical="center", wrap_text=True)

    # 边框
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # 决策标签条件格式颜色
    high_risk_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 红
    medium_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")     # 黄
    low_risk_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # 绿

    return {
        "header_fill": header_fill,
        "header_font": header_font,
        "header_align": header_align,
        "data_font": data_font,
        "data_align": data_align,
        "thin_border": thin_border,
        "high_risk_fill": high_risk_fill,
        "medium_fill": medium_fill,
        "low_risk_fill": low_risk_fill,
    }


def apply_header_style(ws, row: int = 1, col_count: int = 0):
    """应用表头样式到指定工作表。

    Args:
        ws: openpyxl工作表对象
        row: 表头所在行号
        col_count: 列数
    """
    styles = get_excel_styles()
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["header_align"]
        cell.border = styles["thin_border"]


def apply_tag_conditional_format(ws, col_idx: int, row_count: int):
    """对决策标签列应用条件格式（红/黄/绿）。

    Args:
        ws: 工作表对象
        col_idx: 决策标签列序号（1-based）
        row_count: 数据行数
    """
    from openpyxl.formatting.rule import CellIsRule
    styles = get_excel_styles()

    cell_range = f"{ws.cell(row=2, column=col_idx).coordinate}:{ws.cell(row=row_count + 1, column=col_idx).coordinate}"

    # 高=红，中=黄，低=绿
    ws.conditional_formatting.add(cell_range, CellIsRule(
        operator="equal", formula=['"高"'], fill=styles["high_risk_fill"]
    ))
    ws.conditional_formatting.add(cell_range, CellIsRule(
        operator="equal", formula=['"中"'], fill=styles["medium_fill"]
    ))
    ws.conditional_formatting.add(cell_range, CellIsRule(
        operator="equal", formula=['"低"'], fill=styles["low_risk_fill"]
    ))


# ============================================================
# 决策标签推算辅助函数
# ============================================================

def parse_learning_months(learning_duration: str) -> int:
    """从学习时长字符串中解析月份数。

    Args:
        learning_duration: 学习时长字符串，如"8个月""半年""1年"

    Returns:
        月份数（整数），无法解析返回0
    """
    if not learning_duration:
        return 0
    text = learning_duration.strip()
    # 提取数字
    numbers = re.findall(r'\d+', text)
    if numbers:
        n = int(numbers[0])
        if "年" in text:
            return n * 12
        if "半" in text:
            return 6
        return n
    if "半" in text:
        return 6
    return 0


def parse_renewal_count(renewal_history: str) -> int:
    """从续费历史字符串中解析续费次数。

    Args:
        renewal_history: 续费历史字符串，如"续费1次""续过2次""没续过"

    Returns:
        续费次数（整数），无法解析返回0
    """
    if not renewal_history:
        return 0
    text = renewal_history.strip()
    numbers = re.findall(r'\d+', text)
    if numbers:
        return int(numbers[0])
    if "没" in text or "未" in text or "无" in text:
        return 0
    if "续" in text:
        return 1  # "续过"但无具体数字，默认1次
    return 0


def calculate_payment_level(occupation: str, housing_price: str,
                            consumption: str, thresholds: Dict[str, str]) -> str:
    """根据配置阈值推算支付力等级。

    阈值为自然语言描述，这里做关键词匹配。
    workbuddy的AI能力可进一步精确理解这些自然语言规则。

    Args:
        occupation: 家长职业
        housing_price: 小区房价段（AI补齐）
        consumption: 家庭消费力评估（AI补齐）
        thresholds: 支付力阈值配置 {"高": "条件", "中": "条件", "低": "条件"}

    Returns:
        支付力等级："高"/"中"/"低"
    """
    occ = (occupation or "").strip()
    price = (housing_price or "").strip()
    cons = (consumption or "").strip()
    combined = f"{occ} {price} {cons}"

    # 高支付力信号
    high_signals = ["高端", "别墅", "豪宅", "高薪", "高管", "老板", "做生意",
                    "体制内", "公务员", "医生", "律师", "金融", "10万", "8万", "高"]
    # 低支付力信号
    low_signals = ["回迁", "老小区", "城中村", "不稳定", "打工", "临时", "低收入", "低"]

    high_count = sum(1 for s in high_signals if s in combined)
    low_count = sum(1 for s in low_signals if s in combined)

    if high_count >= 1:
        return "高"
    if low_count >= 1:
        return "低"
    return "中"


def calculate_renewal_risk(learning_months: int, class_performance: str,
                           renewal_count: int, thresholds: Dict[str, str]) -> str:
    """根据配置阈值推算续费风险。

    Args:
        learning_months: 学习月份数
        class_performance: 课堂表现
        renewal_count: 续费次数
        thresholds: 续费风险阈值配置

    Returns:
        续费风险："高"/"中"/"低"
    """
    perf = (class_performance or "").strip()

    # 高风险：学习时间短 + 表现下滑 + 没续费
    if learning_months < 3 and renewal_count == 0:
        if any(w in perf for w in ["走神", "下滑", "不专注", "差", "一般"]):
            return "高"
        return "中"

    # 低风险：学习时间长 + 表现稳定 + 已续费
    if learning_months > 6 and renewal_count >= 1:
        if any(w in perf for w in ["专注", "稳定", "好", "积极"]):
            return "低"
        return "中"

    # 中间状态
    if learning_months >= 3 and learning_months <= 6:
        return "中"

    return "中"


def calculate_referral_potential(family_structure: str, plan_goal: str,
                                 ai_awareness: str, thresholds: Dict[str, str]) -> str:
    """根据配置阈值推算转介绍潜力。

    Args:
        family_structure: 家庭结构
        plan_goal: 家长规划目标
        ai_awareness: 对AI认知度
        thresholds: 转介绍潜力阈值配置

    Returns:
        转介绍潜力："高"/"中"/"低"
    """
    score = 0
    struct = (family_structure or "").strip()
    plan = (plan_goal or "").strip()
    ai = (ai_awareness or "").strip()

    # 三代同堂 +1
    if "三代" in struct or "老人" in struct or "爷爷" in struct or "奶奶" in struct:
        score += 1
    # 规划目标明确 +1
    if plan and plan != "不知道" and plan != "不清楚":
        score += 1
    # 对AI认知度高 +1
    if "高" in ai:
        score += 1

    if score >= 2:
        return "高"
    if score == 1:
        return "中"
    return "低"


def calculate_priority(payment: str, risk: str, referral: str,
                       priority_logic: str) -> int:
    """根据配置权重计算跟进优先级（1-5星）。

    Args:
        payment: 支付力等级
        risk: 续费风险
        referral: 转介绍潜力
        priority_logic: 跟进优先级排序逻辑（自然语言描述）

    Returns:
        优先级星级（1-5）
    """
    # 默认权重：支付力0.4 + 续费风险0.3 + 转介绍潜力0.3
    level_map = {"高": 3, "中": 2, "低": 1}

    p_score = level_map.get(payment, 2)
    r_score = level_map.get(risk, 2)
    ref_score = level_map.get(referral, 2)

    # 续费风险高=需要重点跟进（反向）
    risk_weight = 3 if risk == "高" else (2 if risk == "中" else 1)

    weighted = p_score * 0.4 + risk_weight * 0.3 + ref_score * 0.3

    if weighted >= 2.5:
        return 5
    if weighted >= 2.2:
        return 4
    if weighted >= 1.8:
        return 3
    if weighted >= 1.5:
        return 2
    return 1


def match_recommendation(student: Dict[str, Any],
                         rules: List[Dict[str, Any]]) -> Tuple[str, str]:
    """根据推荐规则匹配学生适合的产品方向。

    Args:
        student: 学生记录字典（含所有字段）
        rules: 推荐规则列表

    Returns:
        (推荐课程, 理由) 元组，无匹配返回 ("待配置", "")
    """
    if not rules:
        return ("待配置", "")

    payment = student.get("支付力", "")
    grade = student.get("年级", "")
    interest = student.get("兴趣偏好", "")
    school_tech = student.get("是否科技特色校", "")

    for rule in rules:
        conditions = rule.get("条件", {})
        matched = True

        if "支付力" in conditions and conditions["支付力"]:
            if payment != conditions["支付力"]:
                matched = False

        if matched and "年级" in conditions and conditions["年级"]:
            grade_cond = conditions["年级"]
            if grade_cond == "1-3年级":
                if not any(g in grade for g in ["1年级", "2年级", "3年级"]):
                    matched = False
            elif grade_cond == "4-6年级":
                if not any(g in grade for g in ["4年级", "5年级", "6年级"]):
                    matched = False

        if matched and "兴趣" in conditions and conditions["兴趣"]:
            if "科技" in conditions["兴趣"]:
                if not any(w in interest for w in ["科技", "编程", "电脑", "机器人", "乐高"]):
                    matched = False

        if matched and "学校科技特色" in conditions:
            if conditions["学校科技特色"] is True:
                if "科技特色" not in school_tech and "是" not in school_tech:
                    matched = False

        if matched:
            return (rule.get("推荐课程", "待配置"), rule.get("理由", ""))

    return ("待配置", "")


# ============================================================
# v1.6.0 新增：学情履历与销售漏斗推算函数
# ============================================================

def calculate_current_grade(enrollment_date: str, enrollment_grade: str) -> str:
    """根据入学时间和入学时年级，计算当前年级。

    逻辑：解析入学年级数字（如"3年级"→3），解析入学日期（YYYY-MM），
    计算与当前日期的月份差，每12个月年级+1，返回"X年级"。

    Args:
        enrollment_date: 入学时间，格式 YYYY-MM（如 "2024-09"）
        enrollment_grade: 入学时年级，如 "3年级"

    Returns:
        当前年级字符串，如 "5年级"；输入无效返回空字符串
    """
    if not enrollment_date or not enrollment_grade:
        return ""
    # 提取入学年级数字
    grade_match = re.search(r'(\d+)', enrollment_grade)
    if not grade_match:
        return ""
    base_grade = int(grade_match.group(1))

    # 解析入学日期 YYYY-MM
    date_match = re.match(r'(\d{4})-(\d{1,2})', enrollment_date.strip())
    if not date_match:
        return ""
    enroll_year = int(date_match.group(1))
    enroll_month = int(date_match.group(2))

    now = datetime.now()
    # 计算月份差
    month_diff = (now.year - enroll_year) * 12 + (now.month - enroll_month)
    if month_diff < 0:
        month_diff = 0

    # 每12个月年级+1
    grade_increment = month_diff // 12
    current_grade = base_grade + grade_increment
    return f"{current_grade}年级"


def calculate_enrollment_duration(enrollment_date: str) -> str:
    """根据入学时间计算在读时长。

    逻辑：解析入学日期（YYYY-MM），计算与当前日期的月份差，
    ≥12个月输出"X年Y个月"，<12个月输出"X个月"。

    Args:
        enrollment_date: 入学时间，格式 YYYY-MM

    Returns:
        在读时长字符串，如 "1年2个月""8个月"；输入无效返回空字符串
    """
    if not enrollment_date:
        return ""
    date_match = re.match(r'(\d{4})-(\d{1,2})', enrollment_date.strip())
    if not date_match:
        return ""
    enroll_year = int(date_match.group(1))
    enroll_month = int(date_match.group(2))

    now = datetime.now()
    month_diff = (now.year - enroll_year) * 12 + (now.month - enroll_month)
    if month_diff < 0:
        month_diff = 0

    years = month_diff // 12
    months = month_diff % 12

    if years >= 1:
        if months > 0:
            return f"{years}年{months}个月"
        return f"{years}年"
    return f"{months}个月"


def parse_exam_level(exam_text: str,
                     exam_levels_config: Optional[List[str]] = None) -> str:
    """解析等级考文本，返回等级分类。

    逻辑：从等级考文本（如"机器人3级,Python1级"）中提取最高等级，
    对照配置项判断属于"advanced"(中高级)/"beginner"(初级)/"none"(未考)。
    默认规则：机器人3-4级=中高级、Python2级=中高级、Scratch2级=中高级、其余=初级。

    Args:
        exam_text: 等级考文本，如 "机器人3级,Python1级"
        exam_levels_config: 校区配置中的等级考选项列表（可选）

    Returns:
        "advanced"(中高级) / "beginner"(初级) / "none"(未考)
    """
    text = (exam_text or "").strip()
    if not text or text == "未考":
        return "none"

    # 提取所有"科目+级别"对，用逗号、顿号、空格分隔
    parts = re.split(r'[,，、\s]+', text)
    has_advanced = False
    has_beginner = False

    # 中高级判定关键词映射
    advanced_patterns = {
        "机器人": [3, 4],
        "python": [2],
        "scratch": [2],
    }

    for part in parts:
        part = part.strip()
        if not part:
            continue
        # 提取科目名和级别数字
        level_match = re.search(r'(\d+)', part)
        if not level_match:
            continue
        level_num = int(level_match.group(1))
        subject = part.lower().replace(str(level_num), "").replace("级", "").strip()

        # 匹配中高级
        for adv_subject, adv_levels in advanced_patterns.items():
            if adv_subject in subject and level_num in adv_levels:
                has_advanced = True
                break

        # 任何有效等级都算至少初级
        if level_num >= 1:
            has_beginner = True

    if has_advanced:
        return "advanced"
    if has_beginner:
        return "beginner"
    return "none"


def parse_follow_duration(duration_text: str) -> int:
    """解析累计跟进时长文本为月份数。

    逻辑：支持"3周"→约1个月、"2个月"→2、"半年"→6、"1年"→12等格式。

    Args:
        duration_text: 跟进时长文本

    Returns:
        月份数（整数），无法解析返回0
    """
    if not duration_text:
        return 0
    text = duration_text.strip()

    # 先处理"半年"
    if "半" in text and "年" in text:
        return 6
    if "半年" in text:
        return 6

    numbers = re.findall(r'\d+', text)
    if not numbers:
        return 0

    n = int(numbers[0])

    if "年" in text:
        return n * 12
    if "周" in text:
        # 3周 ≈ 1个月（4周≈1月，向上取整至少1）
        return max(1, round(n / 4))
    if "月" in text:
        return n
    # 纯数字默认按月
    return n


def parse_parent_attention(attention_text: str) -> Dict[str, str]:
    """解析家长关注度多维文本。

    逻辑：支持"群内回复:快|进班:偶尔进|疲态:无"格式解析为结构化字典。
    也支持自然语言模糊匹配。

    Args:
        attention_text: 家长关注度文本

    Returns:
        {"群内回复": "快", "进班": "偶尔进", "疲态": "无"}
        无法解析的字段返回空字符串
    """
    result: Dict[str, str] = {"群内回复": "", "进班": "", "疲态": ""}
    text = (attention_text or "").strip()
    if not text:
        return result

    # 尝试结构化解析：用 | 分隔维度，用 : 或 ：分隔键值
    parts = re.split(r'[|｜]+', text)
    for part in parts:
        part = part.strip()
        if not part:
            continue
        # 匹配 "键:值" 或 "键：值"
        kv_match = re.match(r'(.+?)[：:]\s*(.+)', part)
        if kv_match:
            key = kv_match.group(1).strip()
            val = kv_match.group(2).strip()
            # 模糊匹配键名
            if "群" in key or "回复" in key:
                result["群内回复"] = val
            elif "进班" in key or "进" in key:
                result["进班"] = val
            elif "疲" in key or "累" in key:
                result["疲态"] = val

    # 如果结构化解析全部为空，尝试自然语言模糊匹配
    if not any(result.values()):
        if "回复快" in text or "回复积极" in text or "秒回" in text:
            result["群内回复"] = "快"
        elif "回复慢" in text or "不回复" in text or "无回复" in text:
            result["群内回复"] = "慢"

        if "经常进班" in text or "常进班" in text:
            result["进班"] = "经常进"
        elif "偶尔进班" in text or "偶尔进" in text:
            result["进班"] = "偶尔进"
        elif "从不进班" in text or "不进班" in text:
            result["进班"] = "不进"

        if "疲态" in text or "疲惫" in text:
            if "明显" in text:
                result["疲态"] = "明显"
            elif "轻度" in text or "轻微" in text:
                result["疲态"] = "轻度"
            else:
                result["疲态"] = "有"
        elif "无疲态" in text or "不疲" in text:
            result["疲态"] = "无"

    return result


def calculate_student_profile(student: Dict[str, Any],
                              exam_levels_config: Optional[List[str]] = None) -> List[str]:
    """推算学情画像标签（8类，多标签共存，命中≥3个取前2个）。

    8类标签规则：
    1. 竞赛冲刺型：白名单比赛非空 且 等级考含中高级 且 家长关注度=高
    2. 科创潜力型：特长兴趣(回退兴趣偏好)含"科创/机器人/编程" 且 课堂表现含"积极/主动"
                   且 (白名单比赛为空 或 等级考≤初级)
    3. 兴趣探索型：特长兴趣(回退兴趣偏好)非空 且 在读时长<6个月 且 等级考=未考
    4. 续费稳定型：在读时长≥12个月 且 家长关注度=高 且 疲态=无
    5. 流失风险型：(家长关注度=低 或 群内回复=慢/无回复) 且 疲态=轻度/明显 且 家长新期待为空
    6. 高净值待挖型：(老师侧支付力含"研学/lab/VEX" 或 支付力=高)
                     且 白名单比赛为空 且 (等级考=未考 或 ≤初级)
    7. 谨慎观望型：(累计跟进时长≥2个月 或 对接次数≥3) 且 堵点非空 且 当前阶段≠在读更新
    8. 基础夯实型：在读时长<12个月 且 (等级考≤初级 或 未考) 且 课堂表现含"稳定/踏实" 且 白名单比赛为空

    字段引用说明（v1.6.1修正）：
    - 特长兴趣：优先读D2"特长兴趣"，为空时回退到C表"兴趣偏好"
    - 课堂表现：直接用C表"课堂表现"
    - 家长关注度/家长新期待：从D2学情履历读（已加入汇总表54列）

    推算规则：多标签共存、字段缺失视为不满足、命中≥3个按PROFILE_TAG_PRIORITY排序取前2个。

    Args:
        student: 学生记录字典（含所有汇总表字段）
        exam_levels_config: 校区配置中的等级考选项列表

    Returns:
        学情画像标签列表，如 ["竞赛冲刺型", "续费稳定型"]；无匹配返回空列表
    """
    tags: List[str] = []

    # 提取字段值
    exam_text = (student.get("等级考", "") or "").strip()
    exam_level = parse_exam_level(exam_text, exam_levels_config)
    white_list = (student.get("白名单比赛", "") or "").strip()
    # 特长兴趣：优先用D2的"特长兴趣"，回退到C表的"兴趣偏好"（汇总表有C表字段）
    special_interest = (student.get("特长兴趣", "") or "").strip()
    if not special_interest:
        special_interest = (student.get("兴趣偏好", "") or "").strip()
    # 课堂表现：直接用C表的"课堂表现"（已在汇总表中）
    class_performance = (student.get("课堂表现", "") or "").strip()
    enrollment_duration = (student.get("在读时长", "") or "").strip()
    duration_months = parse_follow_duration(enrollment_duration) if enrollment_duration else parse_learning_months(enrollment_duration)
    parent_attention_raw = (student.get("家长关注度", "") or "").strip()
    parent_attention = parse_parent_attention(parent_attention_raw)
    parent_new_expect = (student.get("家长新期待", "") or "").strip()
    teacher_payment = (student.get("老师侧支付力", "") or "").strip()
    payment_level = (student.get("支付力", "") or "").strip()
    follow_duration = (student.get("累计跟进时长", "") or "").strip()
    follow_months = parse_follow_duration(follow_duration)
    contact_count_raw = student.get("对接次数", "")
    try:
        contact_count = int(contact_count_raw) if contact_count_raw else 0
    except (ValueError, TypeError):
        contact_count = 0
    block_point = (student.get("堵点", "") or "").strip()
    current_stage = (student.get("当前阶段", "") or "").strip()

    # 判断家长关注度高低
    reply_speed = parent_attention.get("群内回复", "")
    fatigue = parent_attention.get("疲态", "")
    attention_high = reply_speed in ("快", "高") or "高" in parent_attention_raw
    attention_low = reply_speed in ("慢", "无", "无回复") or "低" in parent_attention_raw

    # 1. 竞赛冲刺型
    if white_list and exam_level == "advanced" and attention_high:
        tags.append("竞赛冲刺型")

    # 2. 科创潜力型
    sci_keywords = ["科创", "机器人", "编程"]
    active_keywords = ["积极", "主动"]
    has_sci_interest = any(k in special_interest for k in sci_keywords)
    is_active = any(k in class_performance for k in active_keywords)
    if has_sci_interest and is_active and (not white_list or exam_level in ("none", "beginner")):
        tags.append("科创潜力型")

    # 3. 兴趣探索型
    if special_interest and duration_months < 6 and exam_level == "none":
        tags.append("兴趣探索型")

    # 4. 续费稳定型
    if duration_months >= 12 and attention_high and fatigue in ("无", ""):
        tags.append("续费稳定型")

    # 5. 流失风险型
    if (attention_low or reply_speed in ("慢", "无", "无回复")) and fatigue in ("轻度", "明显", "有") and not parent_new_expect:
        tags.append("流失风险型")

    # 6. 高净值待挖型
    wealth_keywords = ["研学", "lab", "VEX", "vex", "Lab"]
    has_wealth_signal = any(k in teacher_payment for k in wealth_keywords) or payment_level == "高"
    if has_wealth_signal and not white_list and exam_level in ("none", "beginner"):
        tags.append("高净值待挖型")

    # 7. 谨慎观望型
    if (follow_months >= 2 or contact_count >= 3) and block_point and current_stage != "在读更新":
        tags.append("谨慎观望型")

    # 8. 基础夯实型
    stable_keywords = ["稳定", "踏实"]
    is_stable = any(k in class_performance for k in stable_keywords)
    if duration_months < 12 and exam_level in ("beginner", "none") and is_stable and not white_list:
        tags.append("基础夯实型")

    # 命中≥3个时按优先级排序取前2个
    if len(tags) >= 3:
        tags.sort(key=lambda t: PROFILE_TAG_PRIORITY.index(t) if t in PROFILE_TAG_PRIORITY else len(PROFILE_TAG_PRIORITY))
        tags = tags[:2]

    return tags


# ============================================================
# v1.7.0 新增：课程产品库匹配与推荐函数
# ============================================================

def match_course_products(profile_tags: List[str],
                          course_library: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """根据学情画像标签匹配课程产品库，返回匹配的课程列表。

    匹配规则：
    - 遍历课程产品库，检查每门课的"适合的学情画像标签"字段
    - 学员画像标签命中课程标签列表中的任一，即视为匹配
    - 多门课程匹配时，按命中标签数降序排序（命中越多优先级越高）
    - 命中数相同时按课程库原始顺序稳定排序
    - 返回Top 3门课程

    Args:
        profile_tags: 学员学情画像标签列表，如["竞赛冲刺型", "续费稳定型"]
        course_library: 课程产品库列表，每项含"适合的学情画像标签"字段

    Returns:
        匹配的课程列表（按适合度排序），每项含课程名+卖点+开班数+匹配标签；
        无匹配返回空列表
    """
    if not profile_tags or not course_library:
        return []

    profile_set = set(profile_tags)
    candidates: List[Tuple[int, int, Dict[str, Any], List[str]]] = []  # (命中数, 原始索引, 课程dict, 命中标签列表)

    for idx, course in enumerate(course_library):
        course_tags = course.get("适合的学情画像标签", [])
        if not isinstance(course_tags, list):
            # 容错：非列表类型尝试转为列表
            course_tags = [course_tags] if course_tags else []
        course_tag_set = set(course_tags)

        # 计算命中标签（交集）
        matched_tags = list(profile_set & course_tag_set)
        if not matched_tags:
            continue

        hit_count = len(matched_tags)
        candidates.append((hit_count, idx, course, matched_tags))

    if not candidates:
        return []

    # 按命中数降序排序，命中数相同按原始索引升序（稳定排序）
    candidates.sort(key=lambda x: (-x[0], x[1]))

    # 取Top 3，构建返回结果（附加"匹配标签"字段）
    result: List[Dict[str, Any]] = []
    for _, _, course, matched_tags in candidates[:3]:
        course_item: Dict[str, Any] = {
            "课程名": course.get("课程名", ""),
            "核心卖点": course.get("核心卖点", []),
            "当前开班数": course.get("当前开班数", 0),
            "匹配标签": matched_tags,
        }
        result.append(course_item)

    return result


def generate_course_recommendation(student: Dict[str, Any],
                                   course_library: List[Dict[str, Any]],
                                   profile_course_map: Dict[str, str]) -> Dict[str, Any]:
    """生成课程推荐（升级版，支持课程级和方向级降级）。

    逻辑：
    1. 调用 calculate_student_profile 获取学情画像标签
    2. 调用 match_course_products 匹配课程产品库
    3. 有匹配 → 返回课程级推荐（课程名+卖点+开班数+话术）
    4. 无匹配/课程库为空 → 降级为 PROFILE_COURSE_MAP 方向级推荐

    Args:
        student: 学员记录（含学情画像相关字段）
        course_library: 课程产品库
        profile_course_map: 画像→方向映射（降级用，即现有的PROFILE_COURSE_MAP）

    Returns:
        {
            "推荐类型": "course_level" / "direction_level",
            "推荐课程": [{"课程名": "...", "核心卖点": [...], "当前开班数": N, "匹配标签": [...]}],  # course_level时
            "推荐方向": "C++/信奥课程",  # direction_level时
            "话术建议": "..."  # course_level时附带，基于第一门匹配课程的卖点生成
        }
    """
    # 1. 获取学情画像标签
    profile_tags = calculate_student_profile(student)

    # 2. 课程库为空 → 降级为方向级推荐
    if not course_library:
        # 取第一个画像标签对应的方向
        direction = "待配置"
        if profile_tags:
            direction = profile_course_map.get(profile_tags[0], "待配置")
        return {
            "推荐类型": "direction_level",
            "推荐方向": direction,
            "话术建议": f"根据画像「{'、'.join(profile_tags) if profile_tags else '未知'}」推荐方向：{direction}",
        }

    # 3. 调用 match_course_products 匹配课程产品库
    matched_courses = match_course_products(profile_tags, course_library)

    # 4. 有匹配 → 课程级推荐
    if matched_courses:
        # 话术建议基于第一门课的核心卖点拼接
        first_course = matched_courses[0]
        selling_points = first_course.get("核心卖点", [])
        if selling_points:
            points_text = "、".join(str(p) for p in selling_points)
            话术建议 = f"推荐「{first_course['课程名']}」，核心优势：{points_text}，当前{first_course.get('当前开班数', 0)}个班在招。"
        else:
            话术建议 = f"推荐「{first_course['课程名']}」，当前{first_course.get('当前开班数', 0)}个班在招。"

        return {
            "推荐类型": "course_level",
            "推荐课程": matched_courses,
            "话术建议": 话术建议,
        }

    # 5. 无匹配 → 降级为方向级推荐
    direction = "待配置"
    if profile_tags:
        direction = profile_course_map.get(profile_tags[0], "待配置")
    return {
        "推荐类型": "direction_level",
        "推荐方向": direction,
        "话术建议": f"课程库中暂无完全匹配课程，根据画像「{'、'.join(profile_tags) if profile_tags else '未知'}」推荐方向：{direction}",
    }


def format_three_way_conflict(consultant_val: str, teacher_val: str,
                              principal_val: str = "") -> str:
    """格式化三路信源冲突为Excel单元格内三值格式。

    Args:
        consultant_val: 顾问录入的值
        teacher_val: 老师补充的值
        principal_val: 校长补充的值（可选，默认空）

    Returns:
        格式化后的字符串，如 "顾问:xxx / 老师:yyy / 校长:zzz [三路待核实]"
        只有两路时："顾问:xxx / 老师:yyy [三路待核实]"
    """
    c_val = (consultant_val or "").strip()
    t_val = (teacher_val or "").strip()
    p_val = (principal_val or "").strip()

    if p_val:
        return f"顾问:{c_val} / 老师:{t_val} / 校长:{p_val} [三路待核实]"
    return f"顾问:{c_val} / 老师:{t_val} [三路待核实]"


# ============================================================
# 脚本结果输出
# ============================================================

def print_script_result(success: bool, message: str, **stats) -> None:
    """输出脚本执行结果（JSON格式到控制台）。

    Args:
        success: 是否成功
        message: 结果消息
        **stats: 统计信息键值对
    """
    result = {
        "success": success,
        "message": message,
    }
    result.update(stats)
    print(json.dumps(result, ensure_ascii=False, indent=2))


# ============================================================
# 模块自测
# ============================================================

if __name__ == "__main__":
    # 测试冲突格式化
    print("=== 冲突格式化测试 ===")
    print(format_conflict_cell("专注", "容易走神"))

    # 测试完成率计算
    print("\n=== 完成率计算测试 ===")
    cache = create_collection_cache("示例校区", "老师", "王老师")
    cache["名单"] = ["张小明", "李小红", "王小刚"]
    cache["已采集记录"] = [{"姓名": "张小明"}]
    print(calculate_completion(cache))

    # 测试学习时长解析
    print("\n=== 学习时长解析测试 ===")
    print(parse_learning_months("8个月"))  # 8
    print(parse_learning_months("半年"))   # 6
    print(parse_learning_months("1年"))    # 12

    # 测试续费次数解析
    print("\n=== 续费次数解析测试 ===")
    print(parse_renewal_count("续费1次"))  # 1
    print(parse_renewal_count("没续过"))   # 0

    # 测试支付力推算
    print("\n=== 支付力推算测试 ===")
    thresholds = {"高": "高端小区", "中": "普通小区", "低": "回迁房"}
    print(calculate_payment_level("做生意", "8-10万/㎡", "高", thresholds))  # 高

    # ===== v1.6.0 新增自测 =====

    # 测试 calculate_current_grade
    print("\n=== calculate_current_grade 测试 ===")
    grade = calculate_current_grade("2024-09", "3年级")
    print(f"入学2024-09, 3年级 → 当前: {grade}")
    assert "年级" in grade, f"应返回含'年级'的字符串, 实际: {grade}"

    # 测试 calculate_enrollment_duration
    print("\n=== calculate_enrollment_duration 测试 ===")
    duration = calculate_enrollment_duration("2024-09")
    print(f"入学2024-09 → 在读时长: {duration}")
    assert duration, f"应返回非空字符串, 实际: {duration}"

    # 测试 parse_exam_level
    print("\n=== parse_exam_level 测试 ===")
    assert parse_exam_level("机器人3级,Python1级") == "advanced", "机器人3级应为advanced"
    assert parse_exam_level("机器人1级") == "beginner", "机器人1级应为beginner"
    assert parse_exam_level("") == "none", "空字符串应为none"
    assert parse_exam_level("未考") == "none", "'未考'应为none"
    assert parse_exam_level("Python2级") == "advanced", "Python2级应为advanced"
    assert parse_exam_level("Scratch2级") == "advanced", "Scratch2级应为advanced"
    assert parse_exam_level("Scratch1级") == "beginner", "Scratch1级应为beginner"
    print("parse_exam_level 全部通过")

    # 测试 parse_follow_duration
    print("\n=== parse_follow_duration 测试 ===")
    assert parse_follow_duration("3周") == 1, f"3周应为1, 实际: {parse_follow_duration('3周')}"
    assert parse_follow_duration("2个月") == 2, f"2个月应为2, 实际: {parse_follow_duration('2个月')}"
    assert parse_follow_duration("1年") == 12, f"1年应为12, 实际: {parse_follow_duration('1年')}"
    assert parse_follow_duration("半年") == 6, f"半年应为6, 实际: {parse_follow_duration('半年')}"
    assert parse_follow_duration("") == 0, "空字符串应为0"
    print("parse_follow_duration 全部通过")

    # 测试 parse_parent_attention
    print("\n=== parse_parent_attention 测试 ===")
    attn = parse_parent_attention("群内回复:快|进班:偶尔进|疲态:无")
    print(f"结构化解析: {attn}")
    assert attn["群内回复"] == "快", f"群内回复应为'快', 实际: {attn['群内回复']}"
    assert attn["进班"] == "偶尔进", f"进班应为'偶尔进', 实际: {attn['进班']}"
    assert attn["疲态"] == "无", f"疲态应为'无', 实际: {attn['疲态']}"
    # 自然语言模糊匹配
    attn2 = parse_parent_attention("群内回复慢，偶尔进班，有轻度疲态")
    print(f"自然语言解析: {attn2}")
    assert attn2["群内回复"] == "慢", f"群内回复应为'慢', 实际: {attn2['群内回复']}"
    print("parse_parent_attention 全部通过")

    # 测试 format_three_way_conflict
    print("\n=== format_three_way_conflict 测试 ===")
    two_way = format_three_way_conflict("A", "B")
    print(f"两路: {two_way}")
    assert "顾问:A" in two_way and "老师:B" in two_way and "[三路待核实]" in two_way
    three_way = format_three_way_conflict("A", "B", "C")
    print(f"三路: {three_way}")
    assert "校长:C" in three_way
    print("format_three_way_conflict 全部通过")

    # 测试 calculate_student_profile — 8类标签各1个命中用例
    print("\n=== calculate_student_profile 测试（8类标签）===")

    # 1. 竞赛冲刺型：白名单比赛非空 + 等级考含中高级 + 家长关注度=高
    s1 = {"等级考": "机器人3级", "白名单比赛": "蓝桥杯2025", "家长关注度": "群内回复:快|进班:经常进|疲态:无"}
    tags1 = calculate_student_profile(s1)
    print(f"竞赛冲刺型: {tags1}")
    assert "竞赛冲刺型" in tags1, f"应含竞赛冲刺型, 实际: {tags1}"

    # 2. 科创潜力型：特长兴趣含科创 + 课堂表现积极 + 无白名单/等级考≤初级
    s2 = {"特长兴趣": "机器人编程", "课堂表现": "积极主动", "等级考": "机器人1级", "白名单比赛": ""}
    tags2 = calculate_student_profile(s2)
    print(f"科创潜力型: {tags2}")
    assert "科创潜力型" in tags2, f"应含科创潜力型, 实际: {tags2}"

    # 3. 兴趣探索型：特长兴趣非空 + 在读时长<6个月 + 等级考=未考
    s3 = {"特长兴趣": "画画", "在读时长": "3个月", "等级考": "", "白名单比赛": "", "课堂表现": ""}
    tags3 = calculate_student_profile(s3)
    print(f"兴趣探索型: {tags3}")
    assert "兴趣探索型" in tags3, f"应含兴趣探索型, 实际: {tags3}"

    # 4. 续费稳定型：在读时长≥12个月 + 家长关注度=高 + 疲态=无
    s4 = {"在读时长": "1年6个月", "家长关注度": "群内回复:快|进班:经常进|疲态:无", "等级考": "", "白名单比赛": "", "特长兴趣": "", "课堂表现": ""}
    tags4 = calculate_student_profile(s4)
    print(f"续费稳定型: {tags4}")
    assert "续费稳定型" in tags4, f"应含续费稳定型, 实际: {tags4}"

    # 5. 流失风险型：家长关注度=低 + 疲态=轻度 + 家长新期待为空
    s5 = {"家长关注度": "群内回复:慢|进班:不进|疲态:轻度", "家长新期待": "", "等级考": "", "白名单比赛": "", "特长兴趣": "", "课堂表现": "", "在读时长": "", "老师侧支付力": "", "支付力": "", "累计跟进时长": "", "对接次数": "", "堵点": "", "当前阶段": ""}
    tags5 = calculate_student_profile(s5)
    print(f"流失风险型: {tags5}")
    assert "流失风险型" in tags5, f"应含流失风险型, 实际: {tags5}"

    # 6. 高净值待挖型：老师侧支付力含研学 + 白名单为空 + 等级考未考
    s6 = {"老师侧支付力": "研学意向强", "白名单比赛": "", "等级考": "", "支付力": "", "特长兴趣": "", "课堂表现": "", "在读时长": "", "家长关注度": "", "家长新期待": "", "累计跟进时长": "", "对接次数": "", "堵点": "", "当前阶段": ""}
    tags6 = calculate_student_profile(s6)
    print(f"高净值待挖型: {tags6}")
    assert "高净值待挖型" in tags6, f"应含高净值待挖型, 实际: {tags6}"

    # 7. 谨慎观望型：累计跟进≥2个月 + 堵点非空 + 当前阶段≠在读更新
    s7 = {"累计跟进时长": "3个月", "对接次数": "", "堵点": "价格犹豫", "当前阶段": "新签挖需", "等级考": "", "白名单比赛": "", "特长兴趣": "", "课堂表现": "", "在读时长": "", "家长关注度": "", "家长新期待": "", "老师侧支付力": "", "支付力": ""}
    tags7 = calculate_student_profile(s7)
    print(f"谨慎观望型: {tags7}")
    assert "谨慎观望型" in tags7, f"应含谨慎观望型, 实际: {tags7}"

    # 8. 基础夯实型：在读时长<12个月 + 等级考≤初级 + 课堂表现含稳定 + 白名单为空
    s8 = {"在读时长": "5个月", "等级考": "机器人1级", "课堂表现": "稳定踏实", "白名单比赛": "", "特长兴趣": "", "家长关注度": "", "家长新期待": "", "老师侧支付力": "", "支付力": "", "累计跟进时长": "", "对接次数": "", "堵点": "", "当前阶段": ""}
    tags8 = calculate_student_profile(s8)
    print(f"基础夯实型: {tags8}")
    assert "基础夯实型" in tags8, f"应含基础夯实型, 实际: {tags8}"

    # 测试命中≥3个取前2个
    print("\n=== calculate_student_profile 多标签截断测试 ===")
    s_multi = {
        "等级考": "机器人3级", "白名单比赛": "蓝桥杯", "家长关注度": "群内回复:快|疲态:无",
        "在读时长": "1年6个月", "特长兴趣": "机器人编程", "课堂表现": "积极主动",
        "累计跟进时长": "3个月", "堵点": "价格犹豫", "当前阶段": "新签挖需",
        "老师侧支付力": "", "支付力": "", "家长新期待": "", "对接次数": "",
    }
    tags_multi = calculate_student_profile(s_multi)
    print(f"多标签命中: {tags_multi}")
    assert len(tags_multi) <= 2, f"命中≥3个应取前2个, 实际: {tags_multi}"
    print("多标签截断通过")

    print("\n=== 列数验证 ===")
    print(f"SUMMARY_COLUMNS 长度: {len(SUMMARY_COLUMNS)}")
    assert len(SUMMARY_COLUMNS) == 59, f"SUMMARY_COLUMNS应为59列, 实际{len(SUMMARY_COLUMNS)}"
    print(f"CONSULTANT_COLUMNS 长度: {len(CONSULTANT_COLUMNS)}")
    assert len(CONSULTANT_COLUMNS) == 28, f"CONSULTANT_COLUMNS应为28列, 实际{len(CONSULTANT_COLUMNS)}"
    print(f"TEACHER_COLUMNS 长度: {len(TEACHER_COLUMNS)}")
    assert len(TEACHER_COLUMNS) == 34, f"TEACHER_COLUMNS应为34列, 实际{len(TEACHER_COLUMNS)}"
    print(f"B2_FIELDS 长度: {len(B2_FIELDS)}")
    assert len(B2_FIELDS) == 9, f"B2_FIELDS应为9字段, 实际{len(B2_FIELDS)}"

    # ===== v1.8.1 新增自测：calculate_credibility =====
    print("\n=== calculate_credibility 测试（4种场景）===")

    # 场景1：低可信度 — flat 字段含不确定关键词
    student_low = {"来源角色": ["顾问"], "冲突标注": ""}
    flat_low = {"家长职业与单位": "好像是在银行工作", "家庭结构": "", "冲突标注": "", "年龄段标签": ""}
    cred_low = calculate_credibility(student_low, flat_low)
    print(f"场景1-低可信度: {cred_low}")
    assert cred_low == "低可信度", f"应返回'低可信度', 实际: {cred_low}"

    # 场景2：待验证 — 冲突标注非空
    student_verify = {"来源角色": ["顾问"], "冲突标注": "冲突字段: 成绩水平"}
    flat_verify = {"家长职业与单位": "医生", "家庭结构": "三口之家",
                   "学校名称": "实验小学", "冲突标注": "冲突字段: 成绩水平",
                   "年龄段标签": "小学中段"}
    cred_verify = calculate_credibility(student_verify, flat_verify)
    print(f"场景2-待验证(冲突): {cred_verify}")
    assert cred_verify == "待验证", f"应返回'待验证', 实际: {cred_verify}"

    # 场景3：高可信度 — 顾问+老师双来源且无冲突
    student_high = {"来源角色": ["顾问", "老师"], "冲突标注": ""}
    flat_high = {"家长职业与单位": "教师/事业单位", "家庭结构": "四口之家",
                 "学校名称": "第一小学", "成绩水平": "优秀",
                 "冲突标注": "", "年龄段标签": "小学高段"}
    cred_high = calculate_credibility(student_high, flat_high)
    print(f"场景3-高可信度: {cred_high}")
    assert cred_high == "高可信度", f"应返回'高可信度', 实际: {cred_high}"

    # 场景4：中可信度 — 默认情况
    student_mid = {"来源角色": ["顾问"], "冲突标注": ""}
    flat_mid = {"家长职业与单位": "个体户", "家庭结构": "三代同堂",
                "学校名称": "中心学校", "冲突标注": "", "年龄段标签": "低龄段"}
    cred_mid = calculate_credibility(student_mid, flat_mid)
    print(f"场景4-中可信度: {cred_mid}")
    assert cred_mid == "中可信度", f"应返回'中可信度', 实际: {cred_mid}"

    # 场景5：待验证 — 所有主观字段为空
    student_empty = {"来源角色": ["顾问"], "冲突标注": ""}
    flat_empty = {"家长职业与单位": "", "家庭结构": "", "教育氛围": "",
                  "居住小区": "", "家长规划目标": "", "家长教育取向": "",
                  "家长竞赛认知": "", "学校名称": "", "成绩水平": "",
                  "性格特点": "", "兴趣偏好": "", "课堂表现": "",
                  "冲突标注": "", "年龄段标签": "小学中段"}
    cred_empty = calculate_credibility(student_empty, flat_empty)
    print(f"场景5-待验证(全空): {cred_empty}")
    assert cred_empty == "待验证", f"应返回'待验证', 实际: {cred_empty}"

    print("calculate_credibility 全部通过 ✓")

    # 验证 create_empty_student_record 新增字段
    print("\n=== create_empty_student_record 验证 ===")
    rec = create_empty_student_record("测试", "示范校区", "顾问")
    assert "销售漏斗" in rec, "record应含'销售漏斗'"
    assert "学情履历" in rec, "record应含'学情履历'"
    assert "家庭背景_老师补充" in rec, "record应含'家庭背景_老师补充'"
    assert len(rec["销售漏斗"]) == 9, f"销售漏斗应有9字段, 实际{len(rec['销售漏斗'])}"
    assert len(rec["学情履历"]) == 13, f"学情履历应有13字段, 实际{len(rec['学情履历'])}"
    assert "顾问侧续费历史" in rec["销售漏斗"], "销售漏斗应含'顾问侧续费历史'"
    print("create_empty_student_record 验证通过")

    # ===== v1.7.0 新增自测：课程产品库匹配与推荐 =====

    # 测试 match_course_products — 命中匹配
    print("\n=== match_course_products 测试 ===")
    test_library_hit = [
        {"课程名": "C++班", "适合的学情画像标签": ["竞赛冲刺型"], "核心卖点": ["竞赛冲刺", "CSP认证"], "当前开班数": 3},
    ]
    result_hit = match_course_products(["竞赛冲刺型"], test_library_hit)
    print(f"命中测试: {result_hit}")
    assert len(result_hit) == 1, f"应返回1门课, 实际{len(result_hit)}"
    assert result_hit[0]["课程名"] == "C++班", f"课程名应为C++班, 实际{result_hit[0]['课程名']}"
    assert "竞赛冲刺型" in result_hit[0]["匹配标签"], f"匹配标签应含竞赛冲刺型"
    print("match_course_products 命中测试通过")

    # 测试 match_course_products — 无匹配返回空列表
    result_miss = match_course_products(["兴趣探索型"], test_library_hit)
    print(f"无匹配测试: {result_miss}")
    assert len(result_miss) == 0, f"无匹配应返回空列表, 实际{len(result_miss)}"
    print("match_course_products 无匹配测试通过")

    # 测试 match_course_products — 多标签命中数排序
    multi_library = [
        {"课程名": "C++班", "适合的学情画像标签": ["竞赛冲刺型"], "核心卖点": ["竞赛"], "当前开班数": 3},
        {"课程名": "综合班", "适合的学情画像标签": ["竞赛冲刺型", "续费稳定型"], "核心卖点": ["竞赛", "续费"], "当前开班数": 2},
    ]
    result_multi = match_course_products(["竞赛冲刺型", "续费稳定型"], multi_library)
    print(f"多标签排序测试: {result_multi}")
    assert len(result_multi) == 2, f"应返回2门课, 实际{len(result_multi)}"
    assert result_multi[0]["课程名"] == "综合班", f"命中数多的应排前, 实际第一门: {result_multi[0]['课程名']}"
    assert len(result_multi[0]["匹配标签"]) == 2, f"综合班应命中2标签"
    print("match_course_products 多标签排序测试通过")

    # 测试 match_course_products — Top 3 截断
    big_library = [
        {"课程名": f"课程{i}", "适合的学情画像标签": ["竞赛冲刺型"], "核心卖点": [f"卖点{i}"], "当前开班数": i}
        for i in range(1, 6)
    ]
    result_top3 = match_course_products(["竞赛冲刺型"], big_library)
    print(f"Top3截断测试: 返回{len(result_top3)}门课")
    assert len(result_top3) == 3, f"应截断为Top 3, 实际{len(result_top3)}"
    print("match_course_products Top3截断测试通过")

    # 测试 generate_course_recommendation — 课程级推荐
    print("\n=== generate_course_recommendation 测试 ===")
    student_course = {
        "等级考": "机器人3级", "白名单比赛": "蓝桥杯2025",
        "家长关注度": "群内回复:快|进班:经常进|疲态:无",
    }
    course_lib = [
        {"课程名": "C++信奥班", "适合的学情画像标签": ["竞赛冲刺型"], "核心卖点": ["CSP-J/S认证", "竞赛冲刺"], "当前开班数": 3},
    ]
    rec_course = generate_course_recommendation(student_course, course_lib, PROFILE_COURSE_MAP)
    print(f"课程级推荐: {rec_course}")
    assert rec_course["推荐类型"] == "course_level", f"应返回course_level, 实际{rec_course['推荐类型']}"
    assert len(rec_course["推荐课程"]) >= 1, f"应有推荐课程"
    assert "话术建议" in rec_course, "应含话术建议"
    print("generate_course_recommendation 课程级测试通过")

    # 测试 generate_course_recommendation — 方向级降级（课程库为空）
    rec_direction_empty = generate_course_recommendation(student_course, [], PROFILE_COURSE_MAP)
    print(f"空库降级: {rec_direction_empty}")
    assert rec_direction_empty["推荐类型"] == "direction_level", f"应返回direction_level, 实际{rec_direction_empty['推荐类型']}"
    assert rec_direction_empty["推荐方向"] == "C++/信奥课程", f"竞赛冲刺型应降级为C++/信奥课程, 实际{rec_direction_empty['推荐方向']}"
    print("generate_course_recommendation 空库降级测试通过")

    # 测试 generate_course_recommendation — 方向级降级（课程库有课但无匹配）
    no_match_lib = [
        {"课程名": "兴趣班", "适合的学情画像标签": ["兴趣探索型"], "核心卖点": ["启蒙"], "当前开班数": 1},
    ]
    rec_direction_nomatch = generate_course_recommendation(student_course, no_match_lib, PROFILE_COURSE_MAP)
    print(f"无匹配降级: {rec_direction_nomatch}")
    assert rec_direction_nomatch["推荐类型"] == "direction_level", f"应返回direction_level, 实际{rec_direction_nomatch['推荐类型']}"
    print("generate_course_recommendation 无匹配降级测试通过")

    # 测试 PROFILE_COURSE_MAP 在 utils.py 中可用
    print("\n=== PROFILE_COURSE_MAP 可用性测试 ===")
    assert PROFILE_COURSE_MAP["竞赛冲刺型"] == "C++/信奥课程", "竞赛冲刺型应映射为C++/信奥课程"
    assert len(PROFILE_COURSE_MAP) == 8, f"PROFILE_COURSE_MAP应有8项, 实际{len(PROFILE_COURSE_MAP)}"
    print("PROFILE_COURSE_MAP 可用性测试通过")

    print("\n[utils.py] 自测全部通过 ✓")
