# -*- coding: utf-8 -*-
"""
校区学员情况管理 Skill - 采集端导出脚本

功能：读取采集端JSON缓存 → 生成个人学员表.xlsx
- 顾问版：含A基础标识 + B1家庭背景 + B2销售漏斗 + E学员细节备注
- 老师版：含A基础标识 + C在校情况 + D1课程成果 + D2学情履历 + E + B_cross_teacher
- 老师版D2.03当前年级/D2.04在读时长自动计算填充
- 含采集完成率sheet

用法：
    python scripts/export_student_xlsx.py --input <JSON缓存路径> --output <xlsx输出路径>

依赖：openpyxl>=3.1.2
"""

import argparse
import json
import os
import sys
from typing import Any, Dict, List

# 添加scripts目录到路径，导入utils
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from utils import (
    A_FIELDS, B_FIELDS, B2_FIELDS, C_FIELDS, D_FIELDS, D2_FIELDS,
    E_FIELDS, CROSS_FIELDS,
    CONSULTANT_COLUMNS, TEACHER_COLUMNS, COMPLETION_COLUMNS,
    get_excel_styles, apply_header_style, calculate_completion,
    calculate_current_grade, calculate_enrollment_duration,
    print_script_result,
)


def flatten_student_record(record: Dict[str, Any], role: str) -> Dict[str, str]:
    """将嵌套的学生记录扁平化为字段名→值的字典。

    顾问版：A基础标识 + B1家庭背景 + B2销售漏斗 + E学员细节备注
    老师版：A基础标识 + C在校情况 + D1课程成果 + D2学情履历 + E + B_cross_teacher
    老师版特殊：D2.03当前年级/D2.04在读时长自动计算填充。

    Args:
        record: 学生记录字典（嵌套结构）
        role: 采集人角色（顾问/老师）

    Returns:
        扁平化后的字段字典
    """
    flat: Dict[str, str] = {}

    # A基础标识字段
    flat["姓名"] = record.get("姓名", "")
    flat["昵称"] = record.get("昵称", "")
    flat["年级"] = record.get("年级", "")
    flat["年龄"] = record.get("年龄", "")
    flat["所在校区"] = record.get("所在校区", record.get("校区", ""))

    if role == "顾问":
        # 顾问版：B1家庭背景字段
        family = record.get("家庭背景", {})
        for f in B_FIELDS:
            flat[f] = family.get(f, "") if family else ""

        # 顾问版：B2销售漏斗字段
        funnel = record.get("销售漏斗", {})
        for f in B2_FIELDS:
            flat[f] = funnel.get(f, "") if funnel else ""
    else:
        # 老师版：C在校情况 + D1课程成果字段
        school = record.get("在校情况", {})
        for f in C_FIELDS:
            flat[f] = school.get(f, "") if school else ""

        course = record.get("课程成果", {})
        for f in D_FIELDS:
            flat[f] = course.get(f, "") if course else ""

        # 老师版：D2学情履历字段
        diary = record.get("学情履历", {})
        for f in D2_FIELDS:
            flat[f] = diary.get(f, "") if diary else ""

        # D2.03当前年级 / D2.04在读时长 自动计算填充（覆盖缓存里的值）
        enroll_date = diary.get("入学时间", "") if diary else ""
        enroll_grade = diary.get("入学时年级", "") if diary else ""
        if enroll_date and enroll_grade:
            flat["当前年级"] = calculate_current_grade(enroll_date, enroll_grade)
        if enroll_date:
            flat["在读时长"] = calculate_enrollment_duration(enroll_date)

        # 老师版：B_cross_teacher 家庭背景(老师补充)
        flat["家庭背景(老师补充)"] = record.get("家庭背景_老师补充", "")

    # E学员细节备注（顾问+老师均可补充）
    for f in E_FIELDS:
        flat[f] = record.get(f, "")

    return flat


def export_xlsx(cache: Dict[str, Any], output_path: str) -> bool:
    """将采集端缓存导出为个人学员表.xlsx。

    Args:
        cache: 采集端缓存字典
        output_path: 输出xlsx文件路径

    Returns:
        导出成功返回True
    """
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    role = cache.get("采集人角色", "老师")
    campus = cache.get("校区", "")
    collector_name = cache.get("采集人姓名", "")
    records = cache.get("已采集记录", [])
    roster = cache.get("名单", [])

    # 确定列顺序
    columns = CONSULTANT_COLUMNS if role == "顾问" else TEACHER_COLUMNS

    # 创建工作簿
    wb = Workbook()

    # ===== Sheet1: 个人学员表 =====
    ws = wb.active
    ws.title = "个人学员表"

    # 写表头
    for col_idx, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    apply_header_style(ws, row=1, col_count=len(columns))

    # 写数据行
    for row_idx, record in enumerate(records, 2):
        flat = flatten_student_record(record, role)
        for col_idx, col_name in enumerate(columns, 1):
            ws.cell(row=row_idx, column=col_idx, value=flat.get(col_name, ""))

    # 设置列宽
    for col_idx in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15

    # 冻结首行
    ws.freeze_panes = "A2"

    # ===== Sheet2: 采集完成率 =====
    ws2 = wb.create_sheet("采集完成率")
    for col_idx, col_name in enumerate(COMPLETION_COLUMNS, 1):
        ws2.cell(row=1, column=col_idx, value=col_name)
    apply_header_style(ws2, row=1, col_count=len(COMPLETION_COLUMNS))

    completion = calculate_completion(cache)
    ws2.cell(row=2, column=1, value=campus)
    ws2.cell(row=2, column=2, value=collector_name)
    ws2.cell(row=2, column=3, value=role)
    ws2.cell(row=2, column=4, value=completion["名单总数"])
    ws2.cell(row=2, column=5, value=completion["已采数"])
    ws2.cell(row=2, column=6, value=completion["完成率"])

    for col_idx in range(1, len(COMPLETION_COLUMNS) + 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 15

    # ===== Sheet3: 待办名单（未采/已采标记） =====
    ws3 = wb.create_sheet("名单进度")
    ws3.cell(row=1, column=1, value="姓名")
    ws3.cell(row=1, column=2, value="状态")
    ws3.cell(row=1, column=3, value="年级")
    apply_header_style(ws3, row=1, col_count=3)

    # 构建已采姓名集合
    collected_names = set()
    for record in records:
        collected_names.add(record.get("姓名", ""))

    row_idx = 2
    for name in roster:
        is_collected = name in collected_names
        # 查找年级
        grade = ""
        for record in records:
            if record.get("姓名") == name:
                grade = record.get("年级", "")
                break
        ws3.cell(row=row_idx, column=1, value=name)
        ws3.cell(row=row_idx, column=2, value="✅已采" if is_collected else "☐未采")
        ws3.cell(row=row_idx, column=3, value=grade)
        row_idx += 1

    ws3.column_dimensions["A"].width = 15
    ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 12

    # 保存文件
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True) if os.path.dirname(output_path) else None
    wb.save(output_path)

    return True


def main():
    """主函数：解析命令行参数并执行导出。"""
    parser = argparse.ArgumentParser(description="采集端：JSON缓存→导出个人学员表.xlsx")
    parser.add_argument("--input", required=True, help="采集端JSON缓存文件路径")
    parser.add_argument("--output", required=True, help="输出xlsx文件路径")
    args = parser.parse_args()

    # 加载JSON缓存
    if not os.path.exists(args.input):
        print_script_result(False, f"输入文件不存在: {args.input}")
        sys.exit(1)

    with open(args.input, "r", encoding="utf-8") as f:
        cache = json.load(f)

    role = cache.get("采集人角色", "老师")
    records = cache.get("已采集记录", [])
    completion = calculate_completion(cache)

    # 导出xlsx
    try:
        success = export_xlsx(cache, args.output)
        if success:
            print_script_result(
                True,
                f"导出成功：{args.output}",
                角色=role,
                已采人数=len(records),
                完成率=completion["完成率"],
            )
        else:
            print_script_result(False, "导出失败")
            sys.exit(1)
    except Exception as e:
        print_script_result(False, f"导出异常: {str(e)}")
        sys.exit(1)


if __name__ == "__main__":
    main()
