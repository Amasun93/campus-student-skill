# -*- coding: utf-8 -*-
"""
校区学员情况管理 Skill - 决策标签推算脚本

功能：读取汇总表 + 校区配置.json → 按配置阈值计算5个决策标签
     + 推算学情画像标签（8类）→ 写入汇总表（条件格式红黄绿）
     → 无匹配规则标"待配置"

决策标签：
1. 支付力等级（高/中/低）- 从配置读支付力阈值
2. 续费风险（高/中/低）- 从配置读续费风险阈值
   v1.7.0：优先D1"续费历史"（老师侧），D1为空回退B2.09"顾问侧续费历史"
3. 转介绍潜力（高/中/低）- 从配置读转介绍潜力阈值
4. 跟进优先级（1-5星）- 从配置读权重逻辑
5. 推荐产品方向 - 从配置读推荐规则匹配

学情画像（8类标签，多标签共存，命中≥3个取前2个）：
竞赛冲刺型/科创潜力型/兴趣探索型/续费稳定型/
流失风险型/高净值待挖型/谨慎观望型/基础夯实型

用法：
    python scripts/write_tags.py --input <汇总表.xlsx> --output <带标签汇总表.xlsx> --config <校区配置.json>

依赖：openpyxl>=3.1.2
"""

import argparse
import json
import os
import sys
from typing import Any, Dict, List, Optional, Tuple

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from utils import (
    A_FIELDS, B_FIELDS, C_FIELDS, D_FIELDS,
    SUMMARY_COLUMNS, AI_FIELDS, TAG_FIELDS,
    PROFILE_TAGS, PROFILE_TAG_PRIORITY,
    get_excel_styles, apply_header_style, apply_tag_conditional_format,
    parse_learning_months, parse_renewal_count,
    calculate_payment_level, calculate_renewal_risk,
    calculate_referral_potential, calculate_priority,
    calculate_student_profile,
    match_recommendation, print_script_result,
    build_course_segment_tags, infer_age_segment,
)


def read_summary_xlsx(file_path: str) -> List[Dict[str, Any]]:
    """读取汇总表.xlsx，解析为学生记录列表（扁平结构）。

    Args:
        file_path: 汇总表xlsx路径

    Returns:
        学生记录列表，每条含所有汇总表列字段
    """
    from openpyxl import load_workbook

    wb = load_workbook(file_path, data_only=True)
    if "学员汇总" not in wb.sheetnames:
        print(f"[错误] 汇总表中未找到'学员汇总'sheet")
        return []

    ws = wb["学员汇总"]
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value) if cell.value else "")

    students: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue

        student: Dict[str, Any] = {}
        for idx, header in enumerate(headers):
            student[header] = str(row[idx]) if idx < len(row) and row[idx] is not None else ""
        students.append(student)

    return students


def calculate_tags_for_student(student: Dict[str, Any],
                               config: Dict[str, Any]) -> Dict[str, Any]:
    """为单个学生计算5个决策标签 + 学情画像标签。

    Args:
        student: 学生记录（扁平结构，含所有字段）
        config: 校区配置字典

    Returns:
        包含5个决策标签和学情画像的字典
    """
    thresholds = config.get("决策标签阈值", {})
    rules = config.get("推荐规则", [])

    # 提取字段值（v1.8.1：家长职业与单位为合并字段）
    occupation = student.get("家长职业与单位", "")
    housing_price = student.get("小区房价段", "")
    consumption = student.get("家庭消费力", "")
    learning_duration = student.get("学习时长", "")
    class_performance = student.get("课堂表现", "")
    renewal_history = student.get("续费历史", "")
    family_structure = student.get("家庭结构", "")
    plan_goal = student.get("家长规划目标", "")
    ai_awareness = student.get("对AI认知度", "")

    # 1. 支付力等级
    payment_thresholds = thresholds.get("支付力", {})
    payment = calculate_payment_level(occupation, housing_price, consumption, payment_thresholds)

    # 2. 续费风险（v1.7.0：优先D1老师侧续费历史，D1为空回退B2.09顾问侧续费历史）
    risk_thresholds = thresholds.get("续费风险", {})
    learning_months = parse_learning_months(learning_duration)
    # D1"续费历史"（老师侧）为第一信源
    renewal_count = parse_renewal_count(renewal_history)
    # D1为空时回退B2.09"顾问侧续费历史"（顾问侧）作补充参考
    if renewal_count == 0:
        consultant_renewal_history = student.get("顾问侧续费历史", "")
        if consultant_renewal_history:
            renewal_count = parse_renewal_count(consultant_renewal_history)
    risk = calculate_renewal_risk(learning_months, class_performance, renewal_count, risk_thresholds)

    # 3. 转介绍潜力
    referral_thresholds = thresholds.get("转介绍潜力", {})
    referral = calculate_referral_potential(family_structure, plan_goal, ai_awareness, referral_thresholds)

    # 4. 跟进优先级
    priority_logic = thresholds.get("跟进优先级", "")
    priority = calculate_priority(payment, risk, referral, priority_logic)

    # 5. 推荐产品方向
    # 构建匹配用的学生字典
    match_student = dict(student)
    match_student["支付力"] = payment
    match_student["是否科技特色校"] = student.get("学校层次(科技特色)", "")
    recommended, reason = match_recommendation(match_student, rules)

    # 6. 学情画像标签（8类，多标签共存，命中≥3个取前2个）
    exam_levels_config = config.get("exam_levels", None)
    # 将支付力写入student副本供profile推算使用
    profile_student = dict(student)
    profile_student["支付力"] = payment
    profile_tags = calculate_student_profile(profile_student, exam_levels_config)

    return {
        "支付力": payment,
        "续费风险": risk,
        "转介绍潜力": referral,
        "跟进优先级": str(priority),
        "推荐产品方向": recommended,
        "推荐理由": reason,
        "学情画像": profile_tags,
    }


def write_tags_xlsx(students: List[Dict[str, Any]],
                    tags_list: List[Dict[str, Any]],
                    template_path: str,
                    output_path: str) -> bool:
    """将带决策标签的数据写入汇总表.xlsx。

    Args:
        students: 学生记录列表
        tags_list: 对应的决策标签列表
        template_path: 原汇总表路径（用于复制其他sheet）
        output_path: 输出文件路径

    Returns:
        写入成功返回True
    """
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    # 基于原汇总表复制
    from openpyxl import Workbook
    wb = load_workbook(template_path)

    # 获取学员汇总sheet
    if "学员汇总" in wb.sheetnames:
        ws = wb["学员汇总"]
    else:
        ws = wb.active

    # 找到决策标签列的索引
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value) if cell.value else "")

    tag_col_map = {}
    for tag in TAG_FIELDS:
        if tag in headers:
            tag_col_map[tag] = headers.index(tag) + 1  # 1-based
        elif tag == "推荐产品方向" and "推荐产品方向" in headers:
            tag_col_map[tag] = headers.index("推荐产品方向") + 1

    # 找到学情画像列索引
    profile_col_idx = None
    if "学情画像" in headers:
        profile_col_idx = headers.index("学情画像") + 1  # 1-based

    # 找到v1.8关系与筛选标签列索引。若旧汇总表无新增列，则保持兼容不新增列。
    relation_col_map = {}
    for relation_col in ["课程段标签", "年龄段标签"]:
        if relation_col in headers:
            relation_col_map[relation_col] = headers.index(relation_col) + 1

    # 写入决策标签和学情画像；同时轻量补算缺失的课程段/年龄段标签。
    for row_idx, (student, tags) in enumerate(zip(students, tags_list), 2):
        if "支付力" in tag_col_map:
            ws.cell(row=row_idx, column=tag_col_map["支付力"], value=tags["支付力"])
        if "续费风险" in tag_col_map:
            ws.cell(row=row_idx, column=tag_col_map["续费风险"], value=tags["续费风险"])
        if "转介绍潜力" in tag_col_map:
            ws.cell(row=row_idx, column=tag_col_map["转介绍潜力"], value=tags["转介绍潜力"])
        if "跟进优先级" in tag_col_map:
            stars = "⭐" * int(tags["跟进优先级"]) if tags["跟进优先级"].isdigit() else tags["跟进优先级"]
            ws.cell(row=row_idx, column=tag_col_map["跟进优先级"], value=stars)
        if "推荐产品方向" in tag_col_map:
            ws.cell(row=row_idx, column=tag_col_map["推荐产品方向"], value=tags["推荐产品方向"])
        # 学情画像（多标签用"、"分隔）
        if profile_col_idx:
            profile_tags = tags.get("学情画像", [])
            profile_text = "、".join(profile_tags) if profile_tags else ""
            ws.cell(row=row_idx, column=profile_col_idx, value=profile_text)

        # v1.8.0 轻量补算课程段标签/年龄段标签，不改变课程推荐核心逻辑。
        if "课程段标签" in relation_col_map:
            current_segment = str(student.get("课程段标签", "") or "").strip()
            if not current_segment:
                course_text = "、".join([
                    str(student.get("在读课程", "") or ""),
                    str(student.get("已报名课程", "") or ""),
                    str(student.get("推荐产品方向", "") or ""),
                ])
                ws.cell(row=row_idx, column=relation_col_map["课程段标签"], value=build_course_segment_tags([], course_text))
        if "年龄段标签" in relation_col_map:
            current_age_segment = str(student.get("年龄段标签", "") or "").strip()
            if not current_age_segment:
                ws.cell(
                    row=row_idx,
                    column=relation_col_map["年龄段标签"],
                    value=infer_age_segment(student.get("年龄", ""), student.get("年级", "")),
                )

    # 应用条件格式
    for tag_col_name in ["支付力", "续费风险", "转介绍潜力"]:
        if tag_col_name in tag_col_map:
            apply_tag_conditional_format(ws, tag_col_map[tag_col_name], len(students))

    # 保存
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True) if os.path.dirname(output_path) else None
    wb.save(output_path)
    return True


def main():
    """主函数：解析参数并执行决策标签推算。"""
    parser = argparse.ArgumentParser(description="决策标签推算：读汇总表+配置→计算标签→写入汇总表")
    parser.add_argument("--input", required=True, help="输入汇总表xlsx路径")
    parser.add_argument("--output", required=True, help="输出带标签汇总表xlsx路径")
    parser.add_argument("--config", required=True, help="校区配置.json路径")
    args = parser.parse_args()

    # 检查文件
    if not os.path.exists(args.input):
        print_script_result(False, f"汇总表不存在: {args.input}")
        sys.exit(1)
    if not os.path.exists(args.config):
        print_script_result(False, f"配置文件不存在: {args.config}")
        sys.exit(1)

    # 加载配置
    with open(args.config, "r", encoding="utf-8") as f:
        config = json.load(f)

    # 读取汇总表
    students = read_summary_xlsx(args.input)
    if not students:
        print_script_result(False, "汇总表为空或读取失败")
        sys.exit(1)

    # 计算决策标签
    tags_list: List[Dict[str, Any]] = []
    for student in students:
        tags = calculate_tags_for_student(student, config)
        tags_list.append(tags)

    # 统计
    high_payment = sum(1 for t in tags_list if t["支付力"] == "高")
    high_risk = sum(1 for t in tags_list if t["续费风险"] == "高")
    high_referral = sum(1 for t in tags_list if t["转介绍潜力"] == "高")
    pending_config = sum(1 for t in tags_list if t["推荐产品方向"] == "待配置")

    # 学情画像分布统计（8类标签各命中几人）
    profile_dist = {}
    for tag in PROFILE_TAGS:
        profile_dist[tag] = sum(1 for t in tags_list if tag in t.get("学情画像", []))
    profile_hit_count = sum(1 for t in tags_list if t.get("学情画像", []))

    # 写入汇总表
    try:
        write_tags_xlsx(students, tags_list, args.input, args.output)
        # 构建学情画像分布输出
        profile_summary = {f"画像_{k}": v for k, v in profile_dist.items()}
        print_script_result(
            True,
            f"标签推算成功：{args.output}",
            总人数=len(students),
            高净值人数=high_payment,
            高续费风险人数=high_risk,
            高转介绍潜力人数=high_referral,
            待配置推荐人数=pending_config,
            学情画像命中人数=profile_hit_count,
            **profile_summary,
        )
    except Exception as e:
        print_script_result(False, f"标签推算异常: {str(e)}")
        sys.exit(1)


if __name__ == "__main__":
    main()
