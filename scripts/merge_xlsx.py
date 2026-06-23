# -*- coding: utf-8 -*-
"""
校区学员情况管理 Skill - 汇总端合并脚本

功能：读取多份个人学员表.xlsx → 按姓名（完整含姓）合并 → 重名用年级对齐
     → A基础标识合并去重 → B1/B2/C/D1/D2冲突检测+A+C双保留 → 输出汇总表.xlsx
     含4个sheet：学员汇总/冲突清单/变更记录/采集完成率

支持增量更新模式：识别新增学生追加 + 已存在学生字段变更更新。

v1.6.0: 新增B2销售漏斗/D2学情履历/家庭背景(老师补充)字段合并，汇总表54列。
v1.7.0: B2.09顾问侧续费历史加入汇总，汇总表55列；D1续费历史(老师侧)+B2.09顾问侧续费历史并列。

用法：
    # 全量合并
    python scripts/merge_xlsx.py --input <文件1.xlsx> <文件2.xlsx> ... --output <汇总表.xlsx>

    # 增量更新
    python scripts/merge_xlsx.py --input <新表.xlsx> --output <更新后.xlsx> --base <已有汇总表.xlsx> --incremental

依赖：openpyxl>=3.1.2
"""

import argparse
import json
import os
import sys
from typing import Any, Dict, List, Optional, Tuple

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from utils import (
    A_FIELDS, B_FIELDS, B2_FIELDS, C_FIELDS, D_FIELDS, D2_FIELDS,
    E_FIELDS, CROSS_FIELDS,
    SUMMARY_COLUMNS, CONFLICT_COLUMNS, CHANGELOG_COLUMNS, COMPLETION_COLUMNS,
    RELATION_MAIN_FIELDS, RESPONSIBILITY_DETAIL_COLUMNS,
    get_excel_styles, apply_header_style, apply_tag_conditional_format,
    format_conflict_cell, format_three_way_conflict,
    create_conflict_record, detect_conflict,
    get_match_key, find_student_in_list, calculate_completion,
    build_relation_snapshot, normalize_course_segment,
    get_timestamp, print_script_result,
    calculate_credibility,
)


def _cell_to_text(value: Any) -> str:
    """将Excel单元格值安全转为字符串。"""
    return str(value).strip() if value is not None else ""


def read_responsibility_details(file_path: str) -> List[Dict[str, str]]:
    """读取个人表或汇总表中的责任关系明细sheet。

    v1.7旧文件没有该sheet时返回空列表，保证合并兼容。

    Args:
        file_path: xlsx文件路径。

    Returns:
        责任关系明细行列表。
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(file_path, data_only=True)
    except Exception:
        return []
    if "责任关系明细" not in wb.sheetnames:
        return []

    ws = wb["责任关系明细"]
    headers = [_cell_to_text(cell.value) for cell in ws[1]]
    details: List[Dict[str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict: Dict[str, str] = {column: "" for column in RESPONSIBILITY_DETAIL_COLUMNS}
        for idx, header in enumerate(headers):
            if not header:
                continue
            row_dict[header] = _cell_to_text(row[idx]) if idx < len(row) else ""
        if not row_dict.get("来源文件"):
            row_dict["来源文件"] = os.path.basename(file_path)
        if any(row_dict.get(column, "") for column in ["姓名", "关系类型", "负责人姓名", "课程段", "关系备注"]):
            details.append(row_dict)
    return details


def dedupe_responsibility_details(details: List[Dict[str, Any]]) -> List[Dict[str, str]]:
    """按规格键去重责任关系明细。

    去重键：姓名+年级+关系类型+关系状态+负责人姓名+课程段+来源文件。
    """
    seen = set()
    result: List[Dict[str, str]] = []
    for detail in details:
        row = {column: _cell_to_text(detail.get(column, "")) for column in RESPONSIBILITY_DETAIL_COLUMNS}
        row["课程段"] = normalize_course_segment(row.get("课程段", ""))
        key = (
            row.get("姓名", ""), row.get("年级", ""), row.get("关系类型", ""),
            row.get("关系状态", ""), row.get("负责人姓名", ""), row.get("课程段", ""),
            row.get("来源文件", ""),
        )
        if key in seen:
            continue
        seen.add(key)
        result.append(row)
    return result


def read_personal_xlsx(file_path: str) -> List[Dict[str, Any]]:
    """读取个人学员表.xlsx，解析为学生记录列表。

    自动检测是顾问版还是老师版（根据列名判断）。
    顾问版解析B2销售漏斗字段，老师版解析D2学情履历字段+B_cross_teacher。

    Args:
        file_path: xlsx文件路径

    Returns:
        学生记录列表，每条记录含采集人角色和所有字段
    """
    from openpyxl import load_workbook

    wb = load_workbook(file_path, data_only=True)

    # 找个人学员表sheet
    ws = None
    for sheet_name in ["个人学员表", "Sheet1", wb.sheetnames[0]]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            break

    if ws is None:
        print(f"[警告] 未找到数据sheet: {file_path}")
        return []

    # 读取表头
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value) if cell.value else "")

    # 判断版本：有"学校名称"列=老师版，有"家长职业"无"学校名称"=顾问版
    is_teacher = "学校名称" in headers
    role = "老师" if is_teacher else "顾问"

    # 读取责任关系明细sheet，按姓名+年级挂到学生记录；v1.7旧文件无sheet时为空。
    detail_rows = read_responsibility_details(file_path)
    detail_map: Dict[Tuple[str, str], List[Dict[str, str]]] = {}
    for detail in detail_rows:
        detail_key = get_match_key(detail.get("姓名", ""), detail.get("年级", ""))
        detail_map.setdefault(detail_key, []).append(detail)

    # 读取数据行
    records: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # 空行跳过
            continue

        row_dict = {}
        for idx, header in enumerate(headers):
            if idx < len(row):
                row_dict[header] = str(row[idx]) if row[idx] is not None else ""
            else:
                row_dict[header] = ""

        # 构建标准记录结构
        record: Dict[str, Any] = {
            "姓名": row_dict.get("姓名", ""),
            "昵称": row_dict.get("昵称", ""),
            "年级": row_dict.get("年级", ""),
            "年龄": row_dict.get("年龄", ""),
            "所在校区": row_dict.get("所在校区", row_dict.get("校区", "")),
            "家庭背景": {},
            "销售漏斗": {},
            "在校情况": {},
            "课程成果": {},
            "学情履历": {},
            "家庭背景_老师补充": "",
            "采集人角色": role,
            "采集来源": os.path.basename(file_path),
        }

        # 填充B1字段（顾问录）
        for f in B_FIELDS:
            record["家庭背景"][f] = row_dict.get(f, "")

        # 填充B2字段（顾问录）
        for f in B2_FIELDS:
            record["销售漏斗"][f] = row_dict.get(f, "")

        # 填充C字段（老师录）
        for f in C_FIELDS:
            record["在校情况"][f] = row_dict.get(f, "")

        # 填充D1字段（老师录）
        for f in D_FIELDS:
            record["课程成果"][f] = row_dict.get(f, "")

        # 填充D2字段（老师录）
        for f in D2_FIELDS:
            record["学情履历"][f] = row_dict.get(f, "")

        # 填充B_cross_teacher（老师版有）
        record["家庭背景_老师补充"] = row_dict.get("家庭背景(老师补充)", "")

        # E学员细节备注
        record["学员细节备注"] = row_dict.get("学员细节备注", "")

        # v1.8.0 主表责任关系快照字段与明细关系
        for f in RELATION_MAIN_FIELDS:
            record[f] = row_dict.get(f, "")
        record_key = get_match_key(record.get("姓名", ""), record.get("年级", ""))
        record["责任关系"] = detail_map.get(record_key, [])

        records.append(record)

    return records


def merge_records(all_records: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """合并所有学生记录，按姓名+年级对齐。

    A基础标识合并去重，B1/B2/C/D1/D2字段冲突检测+A+C双保留。
    B_cross_teacher(家庭背景老师补充)直接填入，不做冲突检测。

    Args:
        all_records: 所有来源的学生记录列表

    Returns:
        (合并后的学生列表, 冲突记录列表)
    """
    merged: Dict[Tuple[str, str], Dict[str, Any]] = {}
    conflicts: List[Dict[str, Any]] = []

    for record in all_records:
        name = record.get("姓名", "").strip()
        grade = record.get("年级", "").strip()
        key = get_match_key(name, grade)

        if not name:
            continue

        if key not in merged:
            # 新学生，直接加入
            merged[key] = {
                "校区": record.get("所在校区", ""),
                "姓名": name,
                "年级": grade,
                "年龄": record.get("年龄", ""),
                "家庭背景": dict(record.get("家庭背景", {})),
                "销售漏斗": dict(record.get("销售漏斗", {})),
                "在校情况": dict(record.get("在校情况", {})),
                "课程成果": dict(record.get("课程成果", {})),
                "学情履历": dict(record.get("学情履历", {})),
                "家庭背景_老师补充": record.get("家庭背景_老师补充", ""),
                "责任关系": list(record.get("责任关系", [])),
                "关系快照": {f: record.get(f, "") for f in RELATION_MAIN_FIELDS},
                "学员细节备注": record.get("学员细节备注", ""),
                "冲突标注": "",
                "来源角色": [record.get("采集人角色", "")],
                "来源文件": [record.get("采集来源", "")],
            }
        else:
            # 已存在学生，合并字段
            existing = merged[key]
            role = record.get("采集人角色", "")

            # 记录来源
            if role not in existing["来源角色"]:
                existing["来源角色"].append(role)

            # 合并A基础标识（去重，取非空值）
            if not existing.get("年龄") and record.get("年龄"):
                existing["年龄"] = record["年龄"]
            if not existing.get("校区") and record.get("所在校区"):
                existing["校区"] = record.get("所在校区")

            # 合并责任关系明细与主表快照字段
            existing_relations = existing.setdefault("责任关系", [])
            existing_relations.extend(record.get("责任关系", []))
            existing["责任关系"] = dedupe_responsibility_details(existing_relations)
            relation_snapshot = existing.setdefault("关系快照", {f: "" for f in RELATION_MAIN_FIELDS})
            for f in RELATION_MAIN_FIELDS:
                new_relation_value = (record.get(f, "") or "").strip()
                if new_relation_value and not relation_snapshot.get(f):
                    relation_snapshot[f] = new_relation_value

            # 合并学员细节备注（自由文本，两边都有的话拼接，不冲突）
            new_detail = (record.get("学员细节备注", "") or "").strip()
            old_detail = (existing.get("学员细节备注", "") or "").strip()
            if new_detail:
                if not old_detail:
                    existing["学员细节备注"] = new_detail
                elif new_detail not in old_detail:
                    # 拼接，标注来源
                    existing["学员细节备注"] = f"[{role}补充] {new_detail} | {old_detail}"

            # 合并B_cross_teacher（老师补充的家庭背景，直接填入，不做冲突检测）
            new_cross = (record.get("家庭背景_老师补充", "") or "").strip()
            if new_cross:
                old_cross = (existing.get("家庭背景_老师补充", "") or "").strip()
                if not old_cross:
                    existing["家庭背景_老师补充"] = new_cross
                elif new_cross not in old_cross:
                    existing["家庭背景_老师补充"] = f"{old_cross} | {new_cross}"

            # 合并B1/B2/C/D1/D2字段，检测冲突
            for group_name, fields, source_data in [
                ("家庭背景", B_FIELDS, record.get("家庭背景", {})),
                ("销售漏斗", B2_FIELDS, record.get("销售漏斗", {})),
                ("在校情况", C_FIELDS, record.get("在校情况", {})),
                ("课程成果", D_FIELDS, record.get("课程成果", {})),
                ("学情履历", D2_FIELDS, record.get("学情履历", {})),
            ]:
                for field in fields:
                    new_val = (source_data.get(field, "") or "").strip()
                    old_val = (existing[group_name].get(field, "") or "").strip()

                    if not new_val:
                        continue  # 新值为空，跳过

                    if not old_val:
                        # 旧值为空，直接填入
                        existing[group_name][field] = new_val
                    elif old_val == new_val:
                        # 值相同，无需处理
                        pass
                    else:
                        # 冲突！A+C双保留
                        # 判断哪个是顾问值，哪个是老师值
                        if role == "顾问":
                            consultant_val, teacher_val = new_val, old_val
                        else:
                            consultant_val, teacher_val = old_val, new_val

                        conflict_cell = format_conflict_cell(consultant_val, teacher_val)
                        existing[group_name][field] = conflict_cell

                        # 记录冲突
                        conflict = create_conflict_record(
                            name, grade, field, consultant_val, teacher_val
                        )
                        conflicts.append(conflict)

                        # 更新冲突标注
                        if existing["冲突标注"]:
                            existing["冲突标注"] += f"; {field}"
                        else:
                            existing["冲突标注"] = f"冲突字段: {field}"

    for student in merged.values():
        student["责任关系"] = dedupe_responsibility_details(student.get("责任关系", []))
        computed_snapshot = build_relation_snapshot(student.get("责任关系", []), student)
        manual_snapshot = student.get("关系快照", {}) if isinstance(student.get("关系快照", {}), dict) else {}
        student["关系快照"] = {
            field: computed_snapshot.get(field, "") or manual_snapshot.get(field, "")
            for field in RELATION_MAIN_FIELDS
        }

    return list(merged.values()), conflicts


def read_completion_from_xlsx(file_path: str) -> Optional[Dict[str, Any]]:
    """从xlsx的采集完成率sheet读取完成率信息。

    Args:
        file_path: xlsx文件路径

    Returns:
        完成率信息字典，无该sheet返回None
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(file_path, data_only=True)
        if "采集完成率" not in wb.sheetnames:
            return None
        ws = wb["采集完成率"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                return {
                    "校区": str(row[0]) if row[0] else "",
                    "采集人": str(row[1]) if row[1] else "",
                    "角色": str(row[2]) if row[2] else "",
                    "名单总数": int(row[3]) if row[3] else 0,
                    "已采数": int(row[4]) if row[4] else 0,
                    "完成率": str(row[5]) if row[5] else "",
                }
    except Exception:
        pass
    return None


def write_summary_xlsx(merged_students: List[Dict[str, Any]],
                       conflicts: List[Dict[str, Any]],
                       completions: List[Dict[str, Any]],
                       changelogs: List[Dict[str, Any]],
                       output_path: str) -> bool:
    """将合并后的数据写入汇总表.xlsx（含5个sheet）。

    Args:
        merged_students: 合并后的学生列表
        conflicts: 冲突记录列表
        completions: 完成率信息列表
        changelogs: 变更记录列表
        output_path: 输出文件路径

    Returns:
        写入成功返回True
    """
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    styles = get_excel_styles()

    # ===== Sheet1: 学员汇总 =====
    ws1 = wb.active
    ws1.title = "学员汇总"

    for col_idx, col_name in enumerate(SUMMARY_COLUMNS, 1):
        ws1.cell(row=1, column=col_idx, value=col_name)
    apply_header_style(ws1, row=1, col_count=len(SUMMARY_COLUMNS))

    for row_idx, student in enumerate(merged_students, 2):
        flat = flatten_merged_student(student)
        for col_idx, col_name in enumerate(SUMMARY_COLUMNS, 1):
            ws1.cell(row=row_idx, column=col_idx, value=flat.get(col_name, ""))

    # 设置列宽
    for col_idx in range(1, len(SUMMARY_COLUMNS) + 1):
        ws1.column_dimensions[get_column_letter(col_idx)].width = 14
    ws1.freeze_panes = "A2"

    # 决策标签列条件格式（支付力/续费风险/转介绍潜力）
    tag_cols = ["支付力", "续费风险", "转介绍潜力"]
    for tag_col in tag_cols:
        if tag_col in SUMMARY_COLUMNS:
            col_idx = SUMMARY_COLUMNS.index(tag_col) + 1
            apply_tag_conditional_format(ws1, col_idx, len(merged_students))

    # ===== Sheet2: 责任关系明细 =====
    ws_rel = wb.create_sheet("责任关系明细")
    for col_idx, col_name in enumerate(RESPONSIBILITY_DETAIL_COLUMNS, 1):
        ws_rel.cell(row=1, column=col_idx, value=col_name)
    apply_header_style(ws_rel, row=1, col_count=len(RESPONSIBILITY_DETAIL_COLUMNS))

    relation_rows: List[Dict[str, str]] = []
    for student in merged_students:
        relation_rows.extend(student.get("责任关系", []))
    relation_rows = dedupe_responsibility_details(relation_rows)
    for row_idx, relation in enumerate(relation_rows, 2):
        for col_idx, col_name in enumerate(RESPONSIBILITY_DETAIL_COLUMNS, 1):
            ws_rel.cell(row=row_idx, column=col_idx, value=relation.get(col_name, ""))

    for col_idx in range(1, len(RESPONSIBILITY_DETAIL_COLUMNS) + 1):
        ws_rel.column_dimensions[get_column_letter(col_idx)].width = 16
    ws_rel.freeze_panes = "A2"

    # ===== Sheet3: 冲突清单 =====
    ws2 = wb.create_sheet("冲突清单")
    for col_idx, col_name in enumerate(CONFLICT_COLUMNS, 1):
        ws2.cell(row=1, column=col_idx, value=col_name)
    apply_header_style(ws2, row=1, col_count=len(CONFLICT_COLUMNS))

    for row_idx, conflict in enumerate(conflicts, 2):
        for col_idx, col_name in enumerate(CONFLICT_COLUMNS, 1):
            ws2.cell(row=row_idx, column=col_idx, value=conflict.get(col_name, ""))

    for col_idx in range(1, len(CONFLICT_COLUMNS) + 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 16

    # ===== Sheet4: 变更记录 =====
    ws3 = wb.create_sheet("变更记录")
    for col_idx, col_name in enumerate(CHANGELOG_COLUMNS, 1):
        ws3.cell(row=1, column=col_idx, value=col_name)
    apply_header_style(ws3, row=1, col_count=len(CHANGELOG_COLUMNS))

    for row_idx, log in enumerate(changelogs, 2):
        for col_idx, col_name in enumerate(CHANGELOG_COLUMNS, 1):
            ws3.cell(row=row_idx, column=col_idx, value=log.get(col_name, ""))

    for col_idx in range(1, len(CHANGELOG_COLUMNS) + 1):
        ws3.column_dimensions[get_column_letter(col_idx)].width = 16

    # ===== Sheet5: 采集完成率 =====
    ws4 = wb.create_sheet("采集完成率")
    for col_idx, col_name in enumerate(COMPLETION_COLUMNS, 1):
        ws4.cell(row=1, column=col_idx, value=col_name)
    apply_header_style(ws4, row=1, col_count=len(COMPLETION_COLUMNS))

    for row_idx, comp in enumerate(completions, 2):
        for col_idx, col_name in enumerate(COMPLETION_COLUMNS, 1):
            ws4.cell(row=row_idx, column=col_idx, value=comp.get(col_name, ""))

    for col_idx in range(1, len(COMPLETION_COLUMNS) + 1):
        ws4.column_dimensions[get_column_letter(col_idx)].width = 15

    # 保存
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True) if os.path.dirname(output_path) else None
    wb.save(output_path)
    return True


def flatten_merged_student(student: Dict[str, Any]) -> Dict[str, str]:
    """将合并后的学生记录扁平化为汇总表列对应的字典（v1.8.1共59列）。

    按SUMMARY_COLUMNS顺序输出，包括：
    - A基础标识(3，v1.8.1移除年龄) + B1家庭背景(7，v1.8.1合并家长职业+单位性质→家长职业与单位，移除对AI认知度)
    + B2销售漏斗(7，含B2.09顾问侧续费历史) + C在校情况(5，v1.8.1移除同伴关系)
    + D1课程成果(5) + D2学情履历(8) + E(1) + 三路信源(1)
    + G责任关系(7) + 冲突标注(1) + 可信度标记(1,v1.8.1新增)
    + AI补齐(7) + 决策标签(5) + 学情画像(1)

    Args:
        student: 合并后的学生记录

    Returns:
        列名→值的字典
    """
    flat: Dict[str, str] = {}

    # A基础标识（3列，v1.8.1移除"年龄"——保留student["年龄"]供年龄段计算，不输出到汇总表）
    flat["校区"] = student.get("校区", "")
    flat["姓名"] = student.get("姓名", "")
    flat["年级"] = student.get("年级", "")

    # B1家庭背景（7列进汇总；v1.8.1合并家长职业+单位性质→家长职业与单位，移除对AI认知度）
    family = student.get("家庭背景", {}) or {}
    # 家长职业与单位合并
    occupation = (family.get("家长职业", "") or "").strip()
    company_type = (family.get("单位性质", "") or "").strip()
    if occupation and company_type:
        flat["家长职业与单位"] = f"{occupation}/{company_type}"
    elif occupation:
        flat["家长职业与单位"] = occupation
    elif company_type:
        flat["家长职业与单位"] = company_type
    else:
        flat["家长职业与单位"] = ""
    # 其余B1字段（排除家长职业、单位性质、对AI认知度）
    b1_summary_fields = ["家庭结构", "教育氛围", "居住小区",
                         "家长规划目标", "家长教育取向", "家长竞赛认知"]
    for f in b1_summary_fields:
        flat[f] = family.get(f, "")

    # B2销售漏斗（7列进汇总；最初兴趣点/介绍过的产品不进汇总；v1.7.0新增B2.09顾问侧续费历史）
    funnel = student.get("销售漏斗", {}) or {}
    b2_summary_fields = ["客户来源", "对接次数", "累计跟进时长", "当前阶段", "堵点", "顾问复盘",
                         "顾问侧续费历史"]
    for f in b2_summary_fields:
        flat[f] = funnel.get(f, "")

    # C在校情况（5列；v1.8.1移除"同伴关系"）
    school = student.get("在校情况", {}) or {}
    c_summary_fields = ["学校名称", "成绩水平", "性格特点", "兴趣偏好", "课堂表现"]
    for f in c_summary_fields:
        flat[f] = school.get(f, "")

    # D1课程成果（5列）
    course = student.get("课程成果", {}) or {}
    for f in D_FIELDS:
        flat[f] = course.get(f, "")

    # D2学情履历（8列进汇总；入学时年级/当前年级/过往奖项/特长兴趣/学生状态观察不进汇总）
    diary = student.get("学情履历", {}) or {}
    d2_summary_fields = ["入学时间", "在读时长", "等级考", "白名单比赛", "老师侧支付力",
                         "家长关注度", "家长新期待", "老师复盘"]
    for f in d2_summary_fields:
        flat[f] = diary.get(f, "")

    # E学员细节备注（1列）
    flat["学员细节备注"] = student.get("学员细节备注", "")

    # 三路信源（1列）
    flat["家庭背景(老师补充)"] = student.get("家庭背景_老师补充", "")

    # G责任关系与筛选标签（7列）
    relation_snapshot = student.get("关系快照", {})
    if not isinstance(relation_snapshot, dict):
        relation_snapshot = {}
    computed_snapshot = build_relation_snapshot(student.get("责任关系", []), flat)
    for f in RELATION_MAIN_FIELDS:
        flat[f] = computed_snapshot.get(f, "") or relation_snapshot.get(f, "") or student.get(f, "") or ""

    # 合并生成（1列）
    flat["冲突标注"] = student.get("冲突标注", "")

    # v1.8.1 可信度标记（1列，新增）
    flat["可信度标记"] = calculate_credibility(student, flat)

    # AI补齐字段（7列，合并阶段为空，后续AI补齐场景填入）
    flat["学校层次(科技特色)"] = ""
    flat["科技特色详情"] = ""
    flat["小区房价段"] = ""
    flat["住户画像"] = ""
    flat["周边竞品"] = ""
    flat["家庭消费力"] = ""
    flat["推荐话术素材"] = ""

    # 决策标签字段（5列，合并阶段为空，后续标签推算场景填入）
    flat["支付力"] = ""
    flat["续费风险"] = ""
    flat["转介绍潜力"] = ""
    flat["跟进优先级"] = ""
    flat["推荐产品方向"] = ""

    # 学情画像（1列，留空，write_tags.py后续填充）
    flat["学情画像"] = ""

    return flat


def incremental_merge(new_records: List[Dict[str, Any]],
                      base_students: List[Dict[str, Any]]
                      ) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
    """增量更新：将新记录合并到已有汇总表中。

    识别新增学生追加，已存在学生字段变更更新（旧值标注）。
    支持B1/B2/C/D1/D2全部字段组的变更检测。

    Args:
        new_records: 新的个人表记录
        base_students: 已有汇总表的学生列表

    Returns:
        (更新后的学生列表, 冲突记录列表, 变更记录列表)
    """
    updated = list(base_students)
    conflicts: List[Dict[str, Any]] = []
    changelogs: List[Dict[str, Any]] = []

    for new_rec in new_records:
        name = new_rec.get("姓名", "").strip()
        grade = new_rec.get("年级", "").strip()
        key = get_match_key(name, grade)

        # 在已有列表中查找
        found_idx = None
        for idx, existing in enumerate(updated):
            if get_match_key(existing.get("姓名", ""), existing.get("年级", "")) == key:
                found_idx = idx
                break

        if found_idx is None:
            # 新增学生，追加
            new_student = {
                "校区": new_rec.get("所在校区", ""),
                "姓名": name,
                "年级": grade,
                "年龄": new_rec.get("年龄", ""),
                "家庭背景": dict(new_rec.get("家庭背景", {})),
                "销售漏斗": dict(new_rec.get("销售漏斗", {})),
                "在校情况": dict(new_rec.get("在校情况", {})),
                "课程成果": dict(new_rec.get("课程成果", {})),
                "学情履历": dict(new_rec.get("学情履历", {})),
                "家庭背景_老师补充": new_rec.get("家庭背景_老师补充", ""),
                "责任关系": list(new_rec.get("责任关系", [])),
                "关系快照": {f: new_rec.get(f, "") for f in RELATION_MAIN_FIELDS},
                "学员细节备注": new_rec.get("学员细节备注", ""),
                "冲突标注": "",
                "来源角色": [new_rec.get("采集人角色", "")],
                "来源文件": [new_rec.get("采集来源", "")],
            }
            updated.append(new_student)

            changelogs.append({
                "姓名": name,
                "变更时间": get_timestamp(),
                "变更字段": "新增学生",
                "旧值": "",
                "新值": "新增",
                "变更来源": new_rec.get("采集来源", ""),
            })
        else:
            # 已存在学生，检测字段变更
            existing = updated[found_idx]
            role = new_rec.get("采集人角色", "")

            for group_name, fields, source_data in [
                ("家庭背景", B_FIELDS, new_rec.get("家庭背景", {})),
                ("销售漏斗", B2_FIELDS, new_rec.get("销售漏斗", {})),
                ("在校情况", C_FIELDS, new_rec.get("在校情况", {})),
                ("课程成果", D_FIELDS, new_rec.get("课程成果", {})),
                ("学情履历", D2_FIELDS, new_rec.get("学情履历", {})),
            ]:
                for field in fields:
                    new_val = (source_data.get(field, "") or "").strip()
                    old_val = (existing[group_name].get(field, "") or "").strip()

                    if not new_val or new_val == old_val:
                        continue

                    if not old_val:
                        existing[group_name][field] = new_val
                        changelogs.append({
                            "姓名": name,
                            "变更时间": get_timestamp(),
                            "变更字段": field,
                            "旧值": "(空)",
                            "新值": new_val,
                            "变更来源": new_rec.get("采集来源", ""),
                        })
                    else:
                        # 字段变更，保留旧值标注
                        conflict_cell = format_conflict_cell(old_val, new_val)
                        existing[group_name][field] = conflict_cell
                        changelogs.append({
                            "姓名": name,
                            "变更时间": get_timestamp(),
                            "变更字段": field,
                            "旧值": old_val,
                            "新值": new_val,
                            "变更来源": new_rec.get("采集来源", ""),
                        })

            # v1.8.0 责任关系明细与主表快照合并
            existing_relations = existing.setdefault("责任关系", [])
            existing_relations.extend(new_rec.get("责任关系", []))
            existing["责任关系"] = dedupe_responsibility_details(existing_relations)
            relation_snapshot = existing.setdefault("关系快照", {f: "" for f in RELATION_MAIN_FIELDS})
            for f in RELATION_MAIN_FIELDS:
                new_relation_value = (new_rec.get(f, "") or "").strip()
                old_relation_value = (relation_snapshot.get(f, "") or "").strip()
                if new_relation_value and new_relation_value != old_relation_value:
                    relation_snapshot[f] = old_relation_value or new_relation_value
                    changelogs.append({
                        "姓名": name,
                        "变更时间": get_timestamp(),
                        "变更字段": f,
                        "旧值": old_relation_value or "(空)",
                        "新值": new_relation_value,
                        "变更来源": new_rec.get("采集来源", ""),
                    })

            # B_cross_teacher 变更检测
            new_cross = (new_rec.get("家庭背景_老师补充", "") or "").strip()
            old_cross = (existing.get("家庭背景_老师补充", "") or "").strip()
            if new_cross and new_cross != old_cross:
                if not old_cross:
                    existing["家庭背景_老师补充"] = new_cross
                    changelogs.append({
                        "姓名": name,
                        "变更时间": get_timestamp(),
                        "变更字段": "家庭背景(老师补充)",
                        "旧值": "(空)",
                        "新值": new_cross,
                        "变更来源": new_rec.get("采集来源", ""),
                    })
                else:
                    existing["家庭背景_老师补充"] = f"{old_cross} | {new_cross}"
                    changelogs.append({
                        "姓名": name,
                        "变更时间": get_timestamp(),
                        "变更字段": "家庭背景(老师补充)",
                        "旧值": old_cross,
                        "新值": new_cross,
                        "变更来源": new_rec.get("采集来源", ""),
                    })

    for student in updated:
        student["责任关系"] = dedupe_responsibility_details(student.get("责任关系", []))
        computed_snapshot = build_relation_snapshot(student.get("责任关系", []), student)
        manual_snapshot = student.get("关系快照", {}) if isinstance(student.get("关系快照", {}), dict) else {}
        student["关系快照"] = {
            field: computed_snapshot.get(field, "") or manual_snapshot.get(field, "") or student.get(field, "") or ""
            for field in RELATION_MAIN_FIELDS
        }

    return updated, conflicts, changelogs


def read_summary_xlsx(file_path: str) -> List[Dict[str, Any]]:
    """读取已有汇总表.xlsx的学生列表（用于增量更新）。

    支持读取B1/B2/C/D1/D2全部字段组及家庭背景(老师补充)。

    Args:
        file_path: 汇总表xlsx路径

    Returns:
        学生记录列表
    """
    from openpyxl import load_workbook

    wb = load_workbook(file_path, data_only=True)
    if "学员汇总" not in wb.sheetnames:
        return []

    summary_details = read_responsibility_details(file_path)
    detail_map: Dict[Tuple[str, str], List[Dict[str, str]]] = {}
    for detail in summary_details:
        detail_key = get_match_key(detail.get("姓名", ""), detail.get("年级", ""))
        detail_map.setdefault(detail_key, []).append(detail)

    ws = wb["学员汇总"]
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value) if cell.value else "")

    students: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue

        row_dict = {}
        for idx, header in enumerate(headers):
            row_dict[header] = str(row[idx]) if idx < len(row) and row[idx] is not None else ""

        student: Dict[str, Any] = {
            "校区": row_dict.get("校区", ""),
            "姓名": row_dict.get("姓名", ""),
            "年级": row_dict.get("年级", ""),
            "年龄": row_dict.get("年龄", ""),
            "家庭背景": {},
            "销售漏斗": {},
            "在校情况": {},
            "课程成果": {},
            "学情履历": {},
            "家庭背景_老师补充": row_dict.get("家庭背景(老师补充)", ""),
            "责任关系": detail_map.get(get_match_key(row_dict.get("姓名", ""), row_dict.get("年级", "")), []),
            "关系快照": {f: row_dict.get(f, "") for f in RELATION_MAIN_FIELDS},
            "学员细节备注": row_dict.get("学员细节备注", ""),
            "冲突标注": row_dict.get("冲突标注", ""),
            "来源角色": [],
            "来源文件": [],
        }

        for f in B_FIELDS:
            student["家庭背景"][f] = row_dict.get(f, "")
        for f in B2_FIELDS:
            student["销售漏斗"][f] = row_dict.get(f, "")
        for f in C_FIELDS:
            student["在校情况"][f] = row_dict.get(f, "")
        for f in D_FIELDS:
            student["课程成果"][f] = row_dict.get(f, "")
        for f in D2_FIELDS:
            student["学情履历"][f] = row_dict.get(f, "")

        students.append(student)

    return students


def main():
    """主函数：解析参数并执行合并/增量更新。"""
    parser = argparse.ArgumentParser(description="汇总端：多表合并+冲突处理+增量更新")
    parser.add_argument("--input", nargs="+", required=True, help="输入的个人学员表xlsx文件路径（可多个）")
    parser.add_argument("--output", required=True, help="输出汇总表xlsx路径")
    parser.add_argument("--base", help="增量更新时的已有汇总表路径")
    parser.add_argument("--incremental", action="store_true", help="增量更新模式")
    args = parser.parse_args()

    # 读取所有输入文件
    all_records: List[Dict[str, Any]] = []
    completions: List[Dict[str, Any]] = []

    for file_path in args.input:
        if not os.path.exists(file_path):
            print(f"[警告] 文件不存在，跳过: {file_path}")
            continue

        records = read_personal_xlsx(file_path)
        all_records.extend(records)

        comp = read_completion_from_xlsx(file_path)
        if comp:
            completions.append(comp)

    if args.incremental and args.base:
        # ===== 增量更新模式 =====
        if not os.path.exists(args.base):
            print_script_result(False, f"基础汇总表不存在: {args.base}")
            sys.exit(1)

        base_students = read_summary_xlsx(args.base)
        updated_students, conflicts, changelogs = incremental_merge(all_records, base_students)

        # 读取基础表的完成率
        base_comp = read_completion_from_xlsx(args.base)
        if base_comp:
            completions.insert(0, base_comp)

        try:
            write_summary_xlsx(updated_students, conflicts, completions, changelogs, args.output)
            new_count = sum(1 for log in changelogs if log["变更字段"] == "新增学生")
            change_count = len(changelogs) - new_count
            print_script_result(
                True,
                f"增量更新成功：{args.output}",
                总人数=len(updated_students),
                新增人数=new_count,
                变更记录数=change_count,
                冲突数=len(conflicts),
            )
        except Exception as e:
            print_script_result(False, f"增量更新异常: {str(e)}")
            sys.exit(1)
    else:
        # ===== 全量合并模式 =====
        merged_students, conflicts = merge_records(all_records)

        try:
            write_summary_xlsx(merged_students, conflicts, completions, [], args.output)
            print_script_result(
                True,
                f"合并成功：{args.output}",
                总人数=len(merged_students),
                冲突数=len(conflicts),
                来源文件数=len(args.input),
            )
        except Exception as e:
            print_script_result(False, f"合并异常: {str(e)}")
            sys.exit(1)


if __name__ == "__main__":
    main()
